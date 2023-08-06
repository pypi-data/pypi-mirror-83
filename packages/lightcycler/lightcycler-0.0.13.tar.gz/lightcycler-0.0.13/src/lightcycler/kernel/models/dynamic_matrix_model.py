import collections
import logging
import os
import re

from PyQt5 import QtCore, QtGui

import openpyxl

import tabula

import numpy as np

import pandas as pd


class InvalidViewError(Exception):
    """Exception raised for dynamic matrix view related errors.
    """


class DynamicMatrixModel(QtCore.QAbstractTableModel):

    propagate_means = QtCore.pyqtSignal(pd.DataFrame)

    def __init__(self, *args, **kwargs):
        """Constructor.
        """

        super(DynamicMatrixModel, self).__init__(*args, **kwargs)

        self._view = 0

        self._dynamic_matrix = pd.DataFrame()
        self._n_values = pd.DataFrame()
        self._means = pd.DataFrame()
        self._stds = pd.DataFrame()

    def on_clear(self):

        self._dynamic_matrix = pd.DataFrame()
        self._n_values = pd.DataFrame()
        self._means = pd.DataFrame()
        self._stds = pd.DataFrame()

        self.layoutChanged.emit()

    def columnCount(self, parent=None):
        """Return the number of columns of the model for a given parent.

        Returns:
            int: the number of columns
        """

        return len(self._n_values.columns)

    def data(self, index, role):
        """Get the data at a given index for a given role.

        Args:
            index (QtCore.QModelIndex): the index
            role (int): the role

        Returns:
            QtCore.QVariant: the data
        """

        if self._n_values.empty:
            return QtCore.QVariant()

        if not index.isValid():
            return QtCore.QVariant()

        row = index.row()
        col = index.column()

        if self._view == 0:
            values = self._means

        elif self._view == 1:
            values = self._stds

        elif self._view == 2:
            values = self._n_values

        else:
            return QtCore.QVariant()

        if role == QtCore.Qt.DisplayRole:
            return str(values.iloc[row, col])

        elif role == QtCore.Qt.BackgroundRole:

            if np.isnan(values.iloc[row, col]):
                return QtGui.QBrush(QtGui.QColor(255, 85, 51))
            else:
                min_value = np.nanmin(values)
                max_value = np.nanmax(values)

                gray_scale = 255 - int(128.0*(values.iloc[row, col] - min_value)/(max_value - min_value))

                return QtGui.QColor(gray_scale, gray_scale, gray_scale)

        elif role == QtCore.Qt.ToolTipRole:

            return str(self._dynamic_matrix.iloc[row, col])

    @property
    def dynamic_matrix(self):
        """Return the dynamic matrix.

        Returns:
            pandas.DataFrame: the dynamic matrix
        """

        return self._dynamic_matrix

    def export(self, workbook):
        """Export the raw data to an excel spreadsheet.

        Args:
            workbook (openpyxl.workbook.workbook.Workbook): the workbook
        """

        workbook.create_sheet('means')
        worksheet = workbook.get_sheet_by_name('means')

        for i, v in enumerate(self._means.columns):
            worksheet.cell(row=1, column=i+2).value = v

        for i, v in enumerate(self._means.index):
            worksheet.cell(row=i+2, column=1).value = v

        for i in range(len(self._means.index)):
            for j in range(len(self._means.columns)):
                worksheet.cell(row=i+2, column=j+2).value = self._means.iloc[i, j]

        workbook.create_sheet('stds')
        worksheet = workbook.get_sheet_by_name('stds')

        for i, v in enumerate(self._stds.columns):
            worksheet.cell(row=1, column=i+2).value = v

        for i, v in enumerate(self._stds.index):
            worksheet.cell(row=i+2, column=1).value = v

        for i in range(len(self._stds.index)):
            for j in range(len(self._stds.columns)):
                worksheet.cell(row=i+2, column=j+2).value = self._stds.iloc[i, j]

        workbook.create_sheet('n_values')
        worksheet = workbook.get_sheet_by_name('n_values')

        for i, v in enumerate(self._n_values.columns):
            worksheet.cell(row=1, column=i+2).value = v

        for i, v in enumerate(self._n_values.index):
            worksheet.cell(row=i+2, column=1).value = v

        for i in range(len(self._n_values.index)):
            for j in range(len(self._n_values.columns)):
                worksheet.cell(row=i+2, column=j+2).value = self._n_values.iloc[i, j]

    def headerData(self, idx, orientation, role):
        """Returns the header data for a given index, orientation and role.

        Args:
            idx (int): the index
            orientation (int): the orientation
            role (int): the role
        """

        if role == QtCore.Qt.DisplayRole:
            if orientation == QtCore.Qt.Horizontal:
                return self._n_values.columns[idx]
            else:
                return self._n_values.index[idx]

    @property
    def means(self):
        """Getter for the means of the dynamic matrix.

        Returns:
            pandas.DataFrame: the means
        """

        return self._means

    @property
    def n_values(self):
        """Return the matrix which stores the number of samples per group.

        Returns:
            pandas.DataFrame: the matrix.
        """

        return self._n_values

    def on_change_value(self, sample, gene, index, new_value):
        """Change a value of the raw data.

            Args:
                sample (str): the selected sample
                gene (str): the selected gene
                index (int): the index
                new_value (float): the new value
        """

        self._means[sample].loc[gene] = round(np.mean(self._dynamic_matrix[sample].loc[gene]), 3)
        self._stds[sample].loc[gene] = round(np.std(self._dynamic_matrix[sample].loc[gene]), 3)
        self._n_values[sample].loc[gene] = len(self._dynamic_matrix[sample].loc[gene])

        self.layoutChanged.emit()

    def on_remove_value(self, sample, gene, index):
        """Remove a value from the dynamic matrix for given sample, genes and index.

        Args:
            sample (str): the selected sample
            gene (str): the selected gene
            index (int): the index
        """

        self._means[sample].loc[gene] = round(np.mean(self._dynamic_matrix[sample].loc[gene]), 3)
        self._stds[sample].loc[gene] = round(np.std(self._dynamic_matrix[sample].loc[gene]), 3)
        self._n_values[sample].loc[gene] = len(self._dynamic_matrix[sample].loc[gene])

        self.layoutChanged.emit()

    def rowCount(self, parent=None):
        """Return the number of rows of the model for a given parent.

        Returns:
            int: the number of rows
        """

        return len(self._n_values.index)

    def set_dynamic_matrix(self, rawdata_model):
        """Build the dynamic matrix from a rawdata model.

        Args:
            rawdata_model (lightcycler.kernel.models.rawdata_model.RawDataModel): the raw data model
        """

        rawdata = rawdata_model.rawdata

        genes = list(collections.OrderedDict.fromkeys(rawdata['Gene']))

        samples = sorted(list(collections.OrderedDict.fromkeys(rawdata['Name'])))

        self._dynamic_matrix = pd.DataFrame(None, index=genes, columns=samples)
        for row in self._dynamic_matrix.index:
            for col in self._dynamic_matrix.columns:
                self._dynamic_matrix.loc[row, col] = []

        for i in range(len(rawdata.index)):
            columns = rawdata.iloc[i]
            gene = columns['Gene']
            sample = columns['Name']
            cp = columns['CP']
            self._dynamic_matrix.loc[gene, sample].append(cp)

        self.update_matrices()

    def update_matrices(self):
        """Update the meand, stdevs and n matrices.
        """

        self._n_values = pd.DataFrame(0, index=self._dynamic_matrix.index, columns=self._dynamic_matrix.columns)
        self._means = pd.DataFrame(np.nan, index=self._dynamic_matrix.index, columns=self._dynamic_matrix.columns)
        self._stds = pd.DataFrame(np.nan, index=self._dynamic_matrix.index, columns=self._dynamic_matrix.columns)

        for i in range(len(self._dynamic_matrix.index)):
            for j in range(len(self._dynamic_matrix.columns)):
                self._n_values.iloc[i, j] = len(self._dynamic_matrix.iloc[i, j])
                if self._dynamic_matrix.iloc[i, j]:
                    self._means.iloc[i, j] = np.mean(self._dynamic_matrix.iloc[i, j])
                    self._stds.iloc[i, j] = np.std(self._dynamic_matrix.iloc[i, j])

        self._means = self._means.round(3)
        self._stds = self._stds.round(3)

        self.layoutChanged.emit()

        self.propagate_means.emit(self._dynamic_matrix)

    def set_view(self, view):

        if view not in range(3):
            raise InvalidViewError('The view must be an integer in [0,2]')

        self._view = view

        self.layoutChanged.emit()

    @property
    def stds(self):
        """Return the standard deviation matrix.

        Returns:
            pandas.DataFrame: the matrix of standard deviations.
        """

        return self._stds
