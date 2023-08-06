import collections
import copy
import csv
import logging
import os
import re

from PyQt5 import QtCore, QtGui

import openpyxl

import tabula

import numpy as np

import pandas as pd


class RawDataError(Exception):
    """This class implements exceptions related with the contents of the data.
    """


class RawDataModel(QtCore.QAbstractTableModel):

    update_dynamic_matrix = QtCore.pyqtSignal(QtCore.QAbstractTableModel)

    def __init__(self, *args, **kwargs):
        """Constructor.
        """

        super(RawDataModel, self).__init__(*args, **kwargs)

        self._rawdata = pd.DataFrame()

        self._rawdata_default = copy.copy(self._rawdata)

    def read_pdf_file(self, pdf_file):
        """Read a PDF data file.

        Args:
            pdf_file: the pdf file
        """

        filename, _ = os.path.splitext(pdf_file)

        basename = os.path.basename(filename)

        match = re.match(r'(\d{4}-\d{2}-\d{2}) .*(RT(1|2|3|1-2))_(\w+)', basename)
        if match is None:
            raise IOError('Invalid filename')

        date, rt, _, gene = match.groups()

        pages = tabula.read_pdf(pdf_file, pages='all')

        data_frame = pd.DataFrame()

        # Loop over the table stored in each page of the pdf document
        for df in pages:
            data_frame = pd.concat([data_frame, df], ignore_index=True)

        # Drop unused columns
        for col in ['Inc', 'Type', 'Concentration', 'Standard', 'Status']:
            data_frame.drop(col, inplace=True, axis=1)

        n_samples = len(data_frame.index)

        # Clean up the Name column from leading "Standard" and "Control" strings
        for i in range(n_samples):
            data_frame['Name'].iloc[i] = ' '.join(data_frame['Name'].iloc[i].split()[1:]).strip()

        data_frame.insert(0, 'Date', [date]*n_samples)

        data_frame.insert(1, 'Gene', [gene]*n_samples)

        data_frame.insert(2, 'RT', [rt]*n_samples)

        data_frame.insert(6, 'File', [basename]*n_samples)

        data_frame['Date'] = pd.to_datetime(data_frame['Date'])
        data_frame['CP'] = data_frame['CP'].str.replace(',', '.').astype(np.float)
        print(data_frame)

        self._rawdata = pd.concat([self._rawdata, data_frame])

    def read_csv_file(self, csv_file):
        """Read a csv data file.

        Args:
            csv_file: the csv file
        """

        filename, ext = os.path.splitext(csv_file)

        basename = os.path.basename(filename)

        match = re.match(r'(\d{4}-\d{2}-\d{2}) .*(RT(1|2|3|1-2))_(\w+)', basename)
        if match is None:
            raise IOError('Invalid filename')

        date, rt, _, gene = match.groups()

        fin = open(csv_file, 'r')
        data = fin.readlines()
        fin.close()

        # Skip the first two lines
        _ = data.pop(0)
        _ = data.pop(0)

        reader = csv.reader(data, delimiter='\t')

        data_frame = pd.DataFrame(columns=['Date', 'Gene', 'RT', 'Pos', 'Name', 'CP', 'File'])

        for row in reader:
            line = {}
            line['Date'] = [date]
            line['Gene'] = [gene]
            line['RT'] = [rt]
            line['Pos'] = [row[2]]
            line['Name'] = [' '.join(row[3].split()[1:]).strip()]
            cp = float(row[4].replace(',', '.')) if row[4].strip() else np.nan
            line['CP'] = [cp]
            line['File'] = [basename]
            data_frame = pd.concat([data_frame, pd.DataFrame.from_dict(line)])

        data_frame['Date'] = pd.to_datetime(data_frame['Date'])

        data_frame.reset_index(drop=True, inplace=True)

        self._rawdata = pd.concat([self._rawdata, data_frame])

    _readers = {'.pdf': read_pdf_file, '.csv': read_csv_file}

    def add_data(self, data_file, sort=False):
        """Add new data to the model.

        Args:
            data (pandas.DataFrame): the data
        """

        _, ext = os.path.splitext(data_file)
        ext = ext.lower()

        try:
            self._readers[ext](self, data_file)
        except KeyError:
            logging.error('Unknown extension for data file {}'.format(data_file))
            return

        if sort:
            self.sort()

        self._rawdata_default = copy.copy(self._rawdata)

        self.layoutChanged.emit()

    def on_clear(self):
        """
        """

        self._rawdata = pd.DataFrame()

        self.layoutChanged.emit()

    def columnCount(self, parent=None):
        """Return the number of columns of the model for a given parent.

        Returns:
            int: the number of columns
        """

        return len(self._rawdata.columns)

    def data(self, index, role):
        """Get the data at a given index for a given role.

        Args:
            index (QtCore.QModelIndex): the index
            role (int): the role

        Returns:
            QtCore.QVariant: the data
        """

        if self._rawdata.empty:
            return QtCore.QVariant()

        if not index.isValid():
            return QtCore.QVariant()

        row = index.row()
        col = index.column()

        if role == QtCore.Qt.DisplayRole:
            return str(self._rawdata.iloc[row, col])
        elif role == QtCore.Qt.ForegroundRole:
            cp_value = self._rawdata['CP'].iloc[row]
            if np.isnan(cp_value):
                return QtGui.QBrush(QtCore.Qt.red)

    def export(self, workbook):
        """Export the raw data to an excel spreadsheet.

        Args:
            workbook (openpyxl.workbook.workbook.Workbook): the workbook
        """

        workbook.create_sheet('raw data')
        worksheet = workbook.get_sheet_by_name('raw data')

        for i, v in enumerate(self._rawdata.columns):
            worksheet.cell(row=1, column=i+1).value = v

        for i in range(len(self._rawdata.index)):
            for j in range(len(self._rawdata.columns)):
                worksheet.cell(row=i+2, column=j+1).value = self._rawdata.iloc[i, j]

    def headerData(self, col, orientation, role):
        """Returns the header data for a given row/column, orientation and role
        """

        if role == QtCore.Qt.DisplayRole:
            if orientation == QtCore.Qt.Horizontal:
                return self._rawdata.columns[col]
            else:
                return str(col+1)
        return None

    def get_row(self, sample, gene, index):

        cond = np.logical_and(self._rawdata['Name'] == sample, self._rawdata['Gene'] == gene)
        matches = self._rawdata.index[cond].tolist()

        if index < 0 or index >= len(matches):
            return None

        return matches[index]

    @ property
    def rawdata(self):
        """Return the raw data.

        Returns:
            pandas.DataFrame: the raw data
        """

        return self._rawdata

    @ rawdata.setter
    def rawdata(self, rawdata):
        """Setter for rawdata.

        Args:
            rawdata (pandas.DataFrame): the raw data
        """

        self._rawdata = rawdata
        self._rawdata_default = copy.copy(self._rawdata)
        self.layoutChanged.emit()

        self.update_dynamic_matrix.emit(self)

    def on_reset(self):

        self._rawdata = copy.copy(self._rawdata_default)

        self.layoutChanged.emit()

        self.update_dynamic_matrix.emit(self)

    def on_change_value(self, sample, gene, index, new_value):
        """Change a value of the raw data.

        Args:
            sample (str): the selected sample
            gene (str): the selected gene
            index (int): the index
            new_value (float): the new value
        """

        cond = np.logical_and(self._rawdata['Name'] == sample, self._rawdata['Gene'] == gene)
        matches = self._rawdata.index[cond].tolist()

        if index < 0 or index >= len(matches):
            return

        self._rawdata.loc[matches[index], 'CP'] = new_value
        self.layoutChanged.emit()

    def on_remove_value(self, sample, gene, index):
        """Remove a value from the dynamic matrix for given sample, genes and index.

        Args:
            sample (str): the selected sample
            gene (str): the selected gene
            index (int): the index
        """

        cond = np.logical_and(self._rawdata['Name'] == sample, self._rawdata['Gene'] == gene)
        matches = self._rawdata.index[cond].tolist()

        if index < 0 or index >= len(matches):
            return

        self.beginRemoveRows(QtCore.QModelIndex(), matches[index], matches[index])

        self._rawdata.drop(matches[index], inplace=True)
        self._rawdata.reset_index(drop=True, inplace=True)

        self.endRemoveRows()

    def rowCount(self, parent=None):
        """Return the number of rows of the model for a given parent.

        Returns:
            int: the number of rows
        """

        return len(self._rawdata.index)

    @ property
    def genes(self):
        """Return the samples names stored in the raw data.

        Returns:
            list of str: the sample names
        """

        genes = list(collections.OrderedDict.fromkeys(self._rawdata['Gene']))

        return genes

    @ property
    def samples(self):
        """Return the samples names stored in the raw data.

        Returns:
            list of str: the sample names
        """

        samples = list(collections.OrderedDict.fromkeys(self._rawdata['Name']))

        return samples

    def sort(self):
        """Sort the raw data.
        """

        if self._rawdata.empty:
            return

        if 'Gene' not in self._rawdata.columns or 'Date' not in self._rawdata.columns:
            raise RawDataError('"Gene" or "Date" columns are missing from the raw data')

        self._rawdata.sort_values(by=['Gene', 'Date'], inplace=True, ascending=[True, True])
