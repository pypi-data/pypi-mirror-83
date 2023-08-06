try:
    import color_utilities
    import static_values
except ModuleNotFoundError:
    from . import color_utilities, static_values
import openpyxl
import math
import jinja2
from typing import Dict, List


def handle_color(
    color: openpyxl.styles.colors.Color,
    themes: List[str],
    alpha: bool=False
):
    """
    handle_color parses the various subtypes of excel colors into a hex-color

    Arguments:
    color: an openpyxl color
    themes: a list of hex-colors defined in the sheet
    alpha: a boolean to determine whether or not to include the alpha channel
    """
    if color is None:
        return None
    if color.type == 'indexed':
        if color.indexed in (64, 65):  # see COLOR_INDEX comment about 64/65
            return None
        color = static_values.COLOR_INDEX[color.indexed]
    elif color.type == 'rgb':
        color = '00' + color_utilities.rgb_and_tint_to_hex(color.rgb, color.tint)
    elif color.type == 'theme':
        if color.theme < len(themes):
            rgb = themes[color.theme]
            color = '00' + color_utilities.rgb_and_tint_to_hex(rgb, color.tint)
        else:
            color = '00000000'
    elif color.type == 'auto':
        color = '00000000'
    else:
        return None

    if not alpha:
        color = color[2:]
    return '#' + color


class ParsedCell:
    def __init__(
        self,
        cell: openpyxl.styles.colors.Color,
        ws_meta: Dict,
        row_idx: int,
        col_idx: int
    ):
        self.text = cell.value or ''
        self.hyperlink = self.handle_hyperlink(cell)
        self.font_style = self.handle_font_style(cell, ws_meta['themes'])
        self.border_style, self.default_border = self.handle_border_style(cell, ws_meta['themes'])
        self.rowspan, self.colspan = self.handle_merged_cells(cell, ws_meta)
        self.row_idx = row_idx
        self.col_idx = col_idx
        self.sizing_style = self.handle_sizing(cell, ws_meta, row_idx, col_idx, self.rowspan, self.colspan)

    @staticmethod
    def handle_hyperlink(cell: openpyxl.styles.colors.Color):
        """
        handle_hyperlink parses the hyperlink target of a cell, if it exists


        Arguments:
        cell: an openpyxl cell

        Returns:
        A string of the hyperlink target, if it exists
        """
        if cell.hyperlink is None:
            return None
        return cell.hyperlink.target

    @staticmethod
    def handle_sizing(
        cell: openpyxl.styles.colors.Color,
        ws_meta: Dict,
        row_idx: int,
        col_idx: int,
        rowspan: int,
        colspan: int
    ):
        """
        handle_sizing uses the cell position and span to figure out the proper width and height of the cell.

        Arguments:
        cell: an openpyxl cell
        ws_meta: A dictionary containing global values for the worksheet (themes, a list of merged_cells)
        row_idx: an integer representing the 0-based row of the cell
        row_idx: an integer representing the 0-based column of the cell
        rowspan: an integer representing the rowspan of the cell
        colspan: an integer representing the colspan of the cell

        Returns:
        a dictionary of alignment and sizing styles
        """
        ret = {}
        width = 0
        for col in range(col_idx, col_idx + colspan):
            width += ws_meta['column_widths'].get(col_idx, ws_meta['default_col_width'])
        height = 0
        for row in range(row_idx, row_idx + rowspan):
            height += ws_meta['row_heights'].get(row_idx, ws_meta['default_row_height'])
        ret['width'] = str(width) + 'px'
        ret['height'] = str(height) + 'px'
        horizontal = cell.alignment.horizontal or 'left'
        vertical = cell.alignment.vertical or 'bottom'
        if vertical == 'center':
            vertical = 'middle'
        ret['text-align'] = horizontal
        ret['vertical-align'] = vertical
        return ret

    @staticmethod
    def handle_merged_cells(
        cell: openpyxl.styles.colors.Color,
        ws_meta: Dict
    ):
        """
        handle_merged_cells uses the view window and the merged_cell_ranges to figure out the proper rowspan and colspan for a cell.

        Arguments:
        cell: an openpyxl cell
        ws_meta: A dictionary containing global values for the worksheet (themes, a list of merged_cells)

        Returns:
        A tuple of two integers representing the (rowspan, colspan) of the cell
        """
        def clamp_to_window(v, direction):
            return max(ws_meta[f'min_{direction}'], min(v, ws_meta[f'max_{direction}']))
        for merge_range in ws_meta['merged_cell_ranges']:
            if cell.coordinate in merge_range:
                rowspan = clamp_to_window(merge_range.max_row, 'row') - clamp_to_window(merge_range.min_row, 'row') + 1
                colspan = clamp_to_window(merge_range.max_col, 'col') - clamp_to_window(merge_range.min_col, 'col') + 1
                return rowspan, colspan
        return 1, 1

    @staticmethod
    def handle_border_style(
        cell: openpyxl.styles.colors.Color,
        themes: List[str]
    ):
        """
        handle_font_style parses all of the styles relating to borders.

        Arguments:
        cell: an openpyxl cell
        themes: a list of hex-colors defined in the sheet

        Returns:
        a dictionary of styles and a dicionary of which borders are default for that cell
        """
        ret = {}
        default_border = {k: False for k in static_values.BORDER_SIDES}
        if (cell.border.top == cell.border.left) and (cell.border.left == cell.border.bottom) and (cell.border.bottom == cell.border.right):  # the borders are all the same
            border = cell.border.top

            border_color = handle_color(border.color, themes) or '#000000'

            border_width = static_values.border_style_to_width.get(border.style)
            border_style = static_values.border_style_to_style.get(border.style, '0px')
            if border_width is not None:
                ret['border'] = f'{border_width} {border_style} {border_color}'
            else:
                ret['border'] = static_values.DEFAULT_BORDER
                default_border = {k: True for k in static_values.BORDER_SIDES}
        else:
            for side in static_values.BORDER_SIDES:
                border = getattr(cell.border, side)

                border_color = handle_color(border.color, themes) or '#000000'

                border_width = static_values.border_style_to_width.get(border.style)
                border_style = static_values.border_style_to_style.get(border.style, '0px')
                if border_width is not None:
                    ret[f'border-{side}'] = f'{border_width} {border_style} {border_color}'
                else:
                    ret[f'border-{side}'] = static_values.DEFAULT_BORDER
                    default_border[side] = True
        return ret, default_border

    @staticmethod
    def handle_font_style(
        cell: openpyxl.styles.colors.Color,
        themes: List[str]
    ):
        """
        handle_font_style parses all of the styles relating to font and background color.

        Arguments:
        cell: an openpyxl cell
        themes: a list of hex-colors defined in the sheet

        Returns:
        a dictionary of styles
        """
        ret = {}
        if cell.font.i:  # italics
            ret['font-style'] = 'italic'
        if cell.font.b:  # bold
            ret['font-weight'] = 'bold'
        if cell.font.u:  # underline
            ret['text-decoration'] = 'underline'
        ret['font-family'] = f"'{cell.font.name}'"
        ret['font-size'] = f"{cell.font.sz}px"

        if cell.fill.patternType is not None:
            background_color = handle_color(cell.fill.fgColor, themes) or handle_color(cell.fill.bgColor, themes)  # foreground color will show above background, right?
            if background_color is not None:
                ret['background-color'] = background_color

        font_color = handle_color(cell.font.color, themes)
        if font_color is not None:
            ret['color'] = font_color
        return ret

    def get_style(self):
        style = []
        for k, v in self.font_style.items():
            style.append(f'{k}: {v}')
        for k, v in self.border_style.items():
            style.append(f'{k}: {v}')
        for k, v in self.sizing_style.items():
            style.append(f'{k}: {v}')
        return '; '.join(style)


def to_html(sheet_cells: List[List[ParsedCell]]):
    """
    Converts the parsed_sheet to an HTML table.

    Arguments:
    sheet_cells: A list of lists of cells, with each inner list representing a row

    Returns:
    a string containing the formatted HTML table
    """
    return jinja2.Template('''
        <table style="border-collapse:collapse">
            {%- for row in sheet_cells -%}
                <tr style="height: {{row[0].height}}">
                    {%- for cell in row -%}
                        {%- if cell.hyperlink is none -%}
                            <td style="{{cell.get_style()}}" rowspan={{cell.rowspan}} colspan={{cell.colspan}}>{{cell.text}}</td>
                        {%- else -%}
                            <td style="{{cell.get_style()}}" rowspan={{cell.rowspan}} colspan={{cell.colspan}}><a href="{{cell.hyperlink}}">{{cell.text}}</a></td>
                        {%- endif -%}
                    {%- endfor -%}
                </tr>
            {%- endfor -%}
        </table>
    ''').render(sheet_cells=sheet_cells, none=None)


def delete_side(cell: ParsedCell, del_side: str):
    """
    Handles the deletion of a border of a cell. If the cell has a "border" style,
    splits it into border-left, border-top, etc.

    Argument:
    cell: a ParsedCell
    del_side: one of ['top', 'right', 'bottom', 'left']

    Returns:
    None
    """
    if cell is None:  # probably a merged cell or the edge of the sheet
        return

    if cell.default_border[del_side] is True:
        if 'border' in cell.border_style:  # we need to convert border to individual border sides to delete one of them
            del cell.border_style['border']
            for side in static_values.BORDER_SIDES:
                cell.border_style[f'border-{side}'] = static_values.DEFAULT_BORDER

        del cell.border_style[f'border-{del_side}']
        cell.default_border[del_side] = False


def fix_borders(sheet_cells: List[List[ParsedCell]], ws_meta: Dict):
    """
    Makes sure that explicitly set borders are not overwritten by default borders.
    Does this by deleting the overlapping default border of any adjcent cell.

    Arguments:
    sheet_cells: A list of lists of cells, with each inner list representing a row
    ws_meta: A dictionary containing global values for the worksheet (themes, a list of merged_cells)

    Returns:
    None
    """
    cell_dict = {}
    for row in sheet_cells:
        for cell in row:
            cell_dict[(cell.row_idx, cell.col_idx)] = cell
    for row in sheet_cells:
        for cell in row:
            for side, is_default in cell.default_border.items():
                if is_default is False:
                    if side == 'top':
                        delete_side(cell_dict.get((cell.row_idx - 1, cell.col_idx)), 'bottom')
                    if side == 'right':
                        delete_side(cell_dict.get((cell.row_idx, cell.col_idx + 1)), 'left')
                    if side == 'bottom':
                        delete_side(cell_dict.get((cell.row_idx + 1, cell.col_idx)), 'top')
                    if side == 'left':
                        delete_side(cell_dict.get((cell.row_idx, cell.col_idx - 1)), 'right')


def fix_background_color(sheet_cells: List[List[ParsedCell]]):
    """
    In an excel, a colored background hides the default border. This function makes the
    same thing happen for the output HTML by deleting all default borders on each cell with
    a background color.

    Arguments:
    sheet_cells: A list of lists of cells, with each inner list representing a row

    Returns:
    None
    """
    for row in sheet_cells:
        for cell in row:
            if cell.font_style.get('background-color') is not None:
                for side, is_default in cell.default_border.items():
                    if is_default:
                        delete_side(cell, side)


def main(
    pathname: str,
    sheetname: str='Sheet1',
    min_row: int=None,
    max_row: int=None,
    min_col: int=None,
    max_col: int=None,
    openpyxl_kwargs: Dict=None
):
    """
    main is the main function. It accepts details about a excel sheet and returns an HTML table matching it.

    Arguments:
    pathname: A path to the excel sheet
    sheetname: The name of the sheet to convert
    min_row: The minimum row to parse in the excel (1-based)
    max_row: The maximum row to parse in the excel (1-based)
    min_col: The minimum column to parse in the excel (1-based)
    max_col: The maximum column to parse in the excel (1-based)
    openpyxl_kwargs: A dicionary of arguments to pass to openpyxl.load_workbook
    """
    def out_of_range(bounds):
        '''bounds are of the form (left_col, top_row, right_col, bottom_row)'''
        return (bounds[0] < (min_col or 0)) or (bounds[1] < (min_row or 0))

    openpyxl_kwargs = openpyxl_kwargs or {}  # just in case people are mutating openpyxl_kwargs between calls.
    wb = openpyxl.load_workbook(pathname, **openpyxl_kwargs)
    ws = wb[sheetname]
    ws_meta = {
        'themes': color_utilities.get_theme_colors(wb),
        'merged_cell_ranges': ws.merged_cells.ranges,
        'column_widths': {(openpyxl.utils.cell.column_index_from_string(i) - 1): math.ceil(x.width * 7) for i, x in ws.column_dimensions.items()},  # converting excel units to pixels
        'default_col_width': ws.sheet_format.defaultColWidth or 64,
        'row_heights': {(i - 1): x.height * (4 / 3) for i, x in ws.row_dimensions.items()},  # converting excel units to pixels
        'default_row_height': ws.sheet_format.defaultRowHeight or 20,
        'min_row': min_row or 1,
        'min_col': min_col or 1,
        'max_row': min(max_row or ws.max_row, ws.max_row),
        'max_col': min(max_col or ws.max_column, ws.max_column),
    }
    parsed_sheet = []
    candidate_merge_ranges = [x for x in ws.merged_cells.ranges if out_of_range(x.bounds)]
    for i, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)):
        parsed_row = []
        for j, cell in enumerate(row):
            if isinstance(cell, openpyxl.cell.cell.Cell):
                parsed_row.append(ParsedCell(cell, ws_meta, i, j))
            else:
                for i, candidate_range in enumerate(candidate_merge_ranges):
                    if cell.coordinate in candidate_range:
                        parent_cell = ws.cell(row=candidate_range.bounds[1], column=candidate_range.bounds[0])
                        parsed_row.append(ParsedCell(parent_cell, ws_meta, i, j))
                        candidate_merge_ranges.pop(i)
                        break
        parsed_sheet.append(parsed_row)

    # it's important to first run background_color and then fix_borders, so that
    # the border can be deleted on the cell in fix_background_color. Then you
    # need to run fix_borders so their neighbors can also have their borders
    # deleted. That's part of the reason we make default_border = False when
    # running delete_side
    fix_background_color(parsed_sheet)
    fix_borders(parsed_sheet, ws_meta)
    body = to_html(parsed_sheet)
    return body
