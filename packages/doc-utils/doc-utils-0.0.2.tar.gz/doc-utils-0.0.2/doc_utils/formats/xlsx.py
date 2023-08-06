import copy

from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from pprint import pprint

from openpyxl import load_workbook
from typing import List, Dict, Union



import re

from doc_utils.formats.base import BaseVariable, BaseDocument

VAR_RX = '\{\s*([a-zA-Z0-9_\-]+):?\s*_*\}'
LOOP_VAR_RX = '\{\s*([a-zA-Z0-9_\-]+)\.\s*([a-zA-Z0-9_\-]+):?\s*_*\}'
FOR_RX = '\{%\s*for\s*([a-zA-Z0-9_\-]+)\s*in\s*([a-zA-Z0-9_\-]+)\s*%\}'
FOR_RX_END = '\{%\s*endfor\s*([a-zA-Z0-9_\-]+)\s*%\}'

class SimpleVariable(BaseVariable):
    def __init__(self, name, expression, cell):
        self.name = name
        self.cell = cell
        self.expression = expression

    def replace(self, value):
        self.cell.value = self.cell.value.replace(self.expression, str(value))


def replace_loop_vars(value, alias, data):
    def _match(m):
        if m.group(1) == alias:
            return data.get(m.group(2),'')

        return m.group(0)

    return re.sub(LOOP_VAR_RX, _match, str(value))



class LoopVariable(BaseVariable):
    def __init__(self, ws, name, alias, start_row):
        self.ws = ws
        self.name = name
        self.alias = alias
        self.start_row = start_row
        self.end_row = None

    def replace(self, values: List[Dict[str, str]]):
        ws = self.ws

        cycle_lines_count = self.end_row - self.start_row + 1
        data_lines_count = cycle_lines_count - 2
        data_values_count = len(values)
        lines_needed = data_lines_count * data_values_count

        shift_needed = lines_needed - cycle_lines_count


        # snapshot data lines
        data = []
        idx = self.start_row + 1
        for row in ws.iter_rows(min_row=self.start_row + 1, max_row=self.end_row - 1):
            cols = []
            for col in row:
                cols.append((
                    col.value or '',
                    copy.copy(col.alignment),
                    copy.copy(col.border),
                    copy.copy(col.fill),
                    copy.copy(col.font)
                ))
            data.append((ws.row_dimensions[idx].height, cols))
            idx += 1

        # add or delete rows as needed
        if shift_needed > 0:
            ws.insert_rows(self.start_row, shift_needed)
        elif shift_needed < 0:
            ws.delete_rows(self.start_row, -shift_needed)

        if shift_needed != 0:
            # shift merged cells up or down as needed
            merged_cells_range = ws.merged_cells.ranges
            for merged_cell in merged_cells_range:
                if merged_cell.max_row > self.start_row:
                    merged_cell.shift(0, shift_needed)

            # shift row heights up or down as needed
            old_heights = []
            idx = self.start_row
            for row in ws.iter_rows(min_row=self.start_row):
                old_heights.append((idx + shift_needed, ws.row_dimensions[idx].height))
                idx += 1
            for idx, height in old_heights:
                ws.row_dimensions[idx].height = height

        # assign values, row height, cell styles
        for value_idx, v in enumerate(values):
            for data_idx, (row_height, data_rows) in enumerate(data):
                row = self.start_row + value_idx * data_lines_count + data_idx
                ws.row_dimensions[row].height = row_height

                for coll_idx, (
                    value,
                    alignment,
                    border,
                    fill,
                    font
                ) in enumerate(data_rows):
                    col = 1 + coll_idx

                    value = replace_loop_vars(value, self.alias, v)

                    cell = ws.cell(row, col)
                    cell.value = value
                    cell.alignment = alignment
                    cell.border = border
                    cell.fill = fill
                    cell.font = font


def find_variables(doc) -> List[Union[SimpleVariable, LoopVariable]]:
    variables = []

    for ws in doc:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not val or not isinstance(val, str):
                    continue

                for match in re.finditer(VAR_RX, val):
                    variables.append(SimpleVariable(
                        match.group(1), match.group(0), cell
                    ))

    loops: Dict[str, LoopVariable] = {}
    for ws in doc:
        for row_nr, row in enumerate(ws.iter_rows()):
            for cell in row:
                val = cell.value
                if not val or not isinstance(val, str):
                    continue

                for_match = re.match(FOR_RX, val)
                if for_match:
                    var_name, name = for_match.groups()
                    v = LoopVariable(ws, name, var_name, row_nr + 1)
                    loops[name] = v

                for_match = re.match(FOR_RX_END, val)
                if for_match:
                    name = for_match.group(1)
                    if name in loops:
                        loops[name].end_row = row_nr + 1
                        variables.append(loops[name])

    return variables


class XlsxDocument(BaseDocument):
    def __init__(self, file_or_path):
        super().__init__(file_or_path)
        self.doc = load_workbook(file_or_path)

    def find_variables(self) -> List[BaseVariable]:
        return find_variables(self.doc)

    def save(self, target):
        self.doc.save(target)
