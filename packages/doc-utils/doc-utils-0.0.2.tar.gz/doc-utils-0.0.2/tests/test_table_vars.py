from openpyxl import Workbook

from doc_utils.formats.xlsx import find_variables, LoopVariable


def print_ws(ws):
    for row in ws.rows:
        for cell in row:
            print(cell, cell.value)

def test_empty():
    doc = Workbook()

    vars = find_variables(doc)

    assert len(vars) == 0


def test_dot_is_not_allowed():
    doc = Workbook()
    ws = doc.active
    ws["A1"] = "A plain {some.cat_name} having some "

    vars = find_variables(doc)

    assert len(vars) == 0


def test_simple_var():
    doc = Workbook()
    ws = doc.active
    ws["A1"] = "A plain {cat_name} having some "

    vars = find_variables(doc)
    assert len(vars) == 1

    vars[0].replace("Puf")
    assert vars[0].name == "cat_name"

    assert ws["A1"].value == 'A plain Puf having some '


def test_simple_loop():
    doc = Workbook()
    ws = doc.active
    ws["A1"] = "{% for x in names %}"
    ws["A2"] = "{x.name}"
    ws["B2"] = "lolo"
    ws["A3"] = "{% endfor names %}"
    assert len(list(ws.rows)) == 3

    vars = find_variables(doc)
    assert len(vars) == 1

    assert isinstance(vars[0], LoopVariable)
    loop = vars[0]  # type: LoopVariable

    assert loop.name == "names"
    assert loop.alias == "x"

    loop.replace([
        {"name": "Puf"}
    ])

    assert ws["A1"].value == 'Puf'
    assert ws["B1"].value == 'lolo'


def test_loop_shift_single():
    doc = Workbook()
    ws = doc.active
    ws["A1"] = "{% for x in names %}"
    ws["A2"] = "{x.name}"
    ws["A3"] = "{% endfor names %}"
    ws["A4"] = "test"

    vars = find_variables(doc)
    vars[0].replace([
        {"name": "Puf"}
    ])

    print_ws(ws)

    # assert len(list(ws.rows)) == 2
    assert ws["A1"].value == 'Puf'
    assert ws["A2"].value == 'test'


def test_loop_shift_two_col():
    doc = Workbook()
    ws = doc.active
    ws["A1"] = "{% for x in names %}"
    ws["A2"] = "{x.name}"
    ws["B2"] = "lala {x.name}"
    ws["A3"] = "{% endfor names %}"
    ws["A4"] = "test"

    vars = find_variables(doc)
    vars[0].replace([
        {"name": "Puf"}
    ])

    assert len(list(ws.rows)) == 2
    assert ws["A1"].value == 'Puf'
    assert ws["B1"].value == 'lala Puf'
    assert ws["A2"].value == 'test'


def test_loop_shift_start():
    doc = Workbook()
    ws = doc.active
    ws["A1"] = "test"
    ws["A2"] = "{% for x in names %}"
    ws["A3"] = "{x.name}"
    ws["A4"] = "{% endfor names %}"

    vars = find_variables(doc)
    vars[0].replace([
        {"name": "Puf"}
    ])

    assert len(list(ws.rows)) == 2
    assert ws["A1"].value == 'test'
    assert ws["A2"].value == 'Puf'


def test_loop_shift_many():
    doc = Workbook()
    ws = doc.active
    ws["A1"] = "{% for x in names %}"
    ws["A2"] = "{x.name}"
    ws["A3"] = "{% endfor names %}"
    ws["A4"] = "test"

    vars = find_variables(doc)
    vars[0].replace([
        {"name": "test1"},
        {"name": "test2"},
        {"name": "test3"},
    ])

    assert len(list(ws.rows)) == 4
    assert ws["A1"].value == 'test1'
    assert ws["A2"].value == 'test2'
    assert ws["A3"].value == 'test3'
    assert ws["A4"].value == 'test'


def test_loop_double():
    doc = Workbook()
    ws = doc.active
    ws["A1"] = "{% for x in names %}"
    ws["A2"] = "{x.name}"
    ws["A3"] = "lala {x.name}"
    ws["A4"] = "{% endfor names %}"
    ws["A5"] = "test"

    vars = find_variables(doc)
    vars[0].replace([
        {"name": "test1"},
        {"name": "test2"},
    ])

    assert len(list(ws.rows)) == 5
    assert ws["A1"].value == 'test1'
    assert ws["A2"].value == 'lala test1'
    assert ws["A3"].value == 'test2'
    assert ws["A4"].value == 'lala test2'
    assert ws["A5"].value == 'test'


def test_loop_double_heights_shift():
    doc = Workbook()
    ws = doc.active
    ws["A1"] = "{% for x in names %}"
    ws["A2"] = "{x.name}"
    ws["A3"] = "lala {x.name}"
    ws["A4"] = "{% endfor names %}"
    ws["A5"] = "test"

    ws.row_dimensions[2].height = 2
    ws.row_dimensions[3].height = 4
    ws.row_dimensions[5].height = 11

    vars = find_variables(doc)
    vars[0].replace([
        {"name": "test1"},
        {"name": "test2"},
        {"name": "test3"},
    ])

    assert len(list(ws.rows)) == 7
    assert ws["A1"].value == 'test1'
    assert ws.row_dimensions[1].height == 2
    assert ws["A2"].value == 'lala test1'
    assert ws.row_dimensions[2].height == 4
    assert ws["A3"].value == 'test2'
    assert ws.row_dimensions[3].height == 2
    assert ws["A4"].value == 'lala test2'
    assert ws.row_dimensions[4].height == 4
    assert ws["A5"].value == 'test3'
    assert ws.row_dimensions[5].height == 2
    assert ws["A6"].value == 'lala test3'
    assert ws.row_dimensions[6].height == 4
    assert ws["A7"].value == 'test'
    assert ws.row_dimensions[7].height == 11



