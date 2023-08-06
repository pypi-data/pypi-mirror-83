"""This module implements the following classes:
    - ExportableDataModel    
"""

import logging

import openpyxl

from PyQt5 import QtCore, QtGui

from pigcel.gui.models.pandas_data_model import PandasDataModel


class ExportableDataModel(PandasDataModel):
    """This class implements a pandas-basd data model which is exportable.
    """

    def export(self, filename):
        """Export the current model to the given excel file.

        Arguments:
            filename (str): the excel filename
        """

        workbook = openpyxl.Workbook()
        # Remove the first empty sheet created by default
        workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))

        workbook.create_sheet('data')

        worksheet = workbook.get_sheet_by_name('data')

        # Write column titles
        for col in range(self.columnCount()):
            col_name = self.headerData(col, QtCore.Qt.Horizontal, role=QtCore.Qt.DisplayRole)
            worksheet.cell(row=1, column=col+2).value = col_name

        # Write row titles
        for row in range(self.rowCount()):
            row_name = self.headerData(row, QtCore.Qt.Vertical, role=QtCore.Qt.DisplayRole)
            worksheet.cell(row=row+2, column=1).value = row_name

        for row in range(self.rowCount()):
            for col in range(self.columnCount()):
                index = self.index(row, col)
                data = self.data(index, QtCore.Qt.DisplayRole)
                worksheet.cell(row=row+2, column=col+2).value = data

        try:
            workbook.save(filename)
        except PermissionError as error:
            logging.error(str(error))
            return
