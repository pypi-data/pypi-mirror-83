import collections
import logging

from PyQt5 import QtCore

import openpyxl

import numpy as np

import pandas as pd

import scipy.stats as stats

import scikit_posthocs as sk

from pigcel.gui.models.animals_pool_model import InvalidPoolData, InvalidTimeError, UnknownPropertyError
from pigcel.kernel.readers.excel_reader import ExcelWorkbookReader
from pigcel.kernel.utils.stats import statistical_functions


class AnimalsGroupsModel(QtCore.QAbstractListModel):
    """This model describes groups of pigs.
    """

    AnimalsPoolModel = QtCore.Qt.UserRole + 1

    Times = QtCore.Qt.UserRole + 2

    def __init__(self, parent):
        """Constructor.

        Args:
            parent (PyQt5.QtWidgets.QObject): the parent object
        """

        super(AnimalsGroupsModel, self).__init__(parent)

        self._groups = collections.OrderedDict()

    def add_model(self, name, model):
        """Add a new monitoring pool model to this model.

        Args:
            name (str): the name of the model to add
            model (pigcel.gui.models.animals_pool_model.AnimalsPoolModel): the model
        """

        if name in self._groups:
            return

        self.beginInsertRows(QtCore.QModelIndex(), self.rowCount(), self.rowCount())

        self._groups[name] = [model, True]

        self.endInsertRows()

    def data(self, index, role):
        """Return the data for a given index and a given role

        Args:
            index (QtCore.QModelIndex): the index
            role (int): the role

        Returns:
            QtCore.QVariant: the data
        """

        if not index.isValid():
            return QtCore.QVariant()

        row = index.row()

        group_names = list(self._groups.keys())

        group = group_names[row]

        if role == QtCore.Qt.DisplayRole:
            return group
        elif role == QtCore.Qt.CheckStateRole:
            return QtCore.Qt.Checked if self._groups[group][1] else QtCore.Qt.Unchecked
        elif role == AnimalsGroupsModel.AnimalsPoolModel:
            return self._groups[group][0]
        else:
            return QtCore.QVariant()

    def evaluate_global_group_effect(self, selected_property):
        """Evaluate the group effect for a given property for the selected groups.

        Args:
            selected_property (str): the selected property

        Returns:
            pandas.DataFrame: the mann-whitney or kruskal_wallis statistics for each time. The n first columns gives the number of animals found in the n groups 
            and the n+1 th value gives the p value
        """

        selected_groups = [(k, v[0]) for k, v in self._groups.items() if v[1]]
        if len(selected_groups) < 2:
            logging.error('Less than two groups selected for evaluating global group effect')
            return None

        data_pooled_per_time = collections.OrderedDict()

        # Retrieve all the times defined across all the pool data
        all_pool_data = []
        all_times = set()
        for _, animals_pool_model in selected_groups:
            all_pool_data.append(animals_pool_model.get_pool_data(selected_property))
            all_times.update(all_pool_data[-1].index)
        all_times = sorted(all_times)

        # Loop over  all the pool data
        for pool_data in all_pool_data:
            for time in all_times:
                # If time is a defined in the pool_data, filter out the corresponding row from any nan value
                if time in pool_data.index:
                    nan_filtered_values = [v for v in pool_data.loc[time] if not np.isnan(v)]
                else:
                    nan_filtered_values = []
                data_pooled_per_time.setdefault(time, []).append(nan_filtered_values)

        # p_values_per_time will be a DataFrame whose index are the times and the n - 1 first columns are the number of animal for each group
        # while the nth value is the p value
        p_values_per_time = []
        for values_per_group in data_pooled_per_time.values():
            # If the number of group is < 2, the group effect can not be assessed. Set the p value to nan
            if len(values_per_group) < 2:
                p_value = np.nan
            else:
                # If one of the group has no valid values, set the p value to nan
                if [] in values_per_group:
                    p_value = np.nan
                else:
                    try:
                        if len(values_per_group) > 2:
                            p_value = stats.kruskal(*values_per_group).pvalue
                        else:
                            p_value = stats.mannwhitneyu(*values_per_group, alternative='two-sided').pvalue
                    except ValueError as error:
                        p_value = np.nan
                        logging.warning(str(error))
            p_values_per_time.append([len(v) for v in values_per_group] + [p_value])

        group_names = [v[0] for v in selected_groups]
        columns = group_names + ['p value']
        p_values = pd.DataFrame(p_values_per_time, index=list(data_pooled_per_time.keys()), columns=columns)

        return p_values

    def evaluate_global_time_effect(self, selected_property):
        """Evaluate the time effect for a given property for the selected groups.

        Args:
            selected_property (str): the selected property

        Returns:
            dict: the time status for each selected group
            pandas.DataFrame: the Friedman chi square p_values for the selected groups
        """

        selected_groups = [(k, v[0]) for k, v in self._groups.items() if v[1]]
        if len(selected_groups) < 1:
            logging.error('No group selected for evaluating the global time effect')
            return

        p_values = []
        group_names = []
        times_per_group = {}
        # Loop over the animal pools
        for group, animals_pool_model in selected_groups:
            times, p_value = animals_pool_model.evaluate_global_time_effect(selected_property)
            p_values.append(p_value)
            group_names.append(group)
            times_per_group[group] = times

        p_values = pd.DataFrame(p_values, index=group_names, columns=['p-value'])

        return times_per_group, p_values

    def evaluate_pairwise_group_effect(self, selected_property):
        """Evaluate the group effect pairwisely for a given property for the selected groups.

        Args:
            selected_property (str): the selected property

        Returns:
            collections.OrderedDict: Dunn p-values matrix stored as a pandas.DataFrame for each time
        """

        selected_groups = [(k, v[0]) for k, v in self._groups.items() if v[1]]
        if len(selected_groups) < 2:
            logging.error('Less than two groups selected for evaluating global group effect')
            return

        data_pooled_per_time = collections.OrderedDict()

        # Take all the times defined across all the pool data
        all_pool_data = []
        all_times = set()
        for _, animals_pool_model in selected_groups:
            all_pool_data.append(animals_pool_model.get_pool_data(selected_property))
            all_times.update(all_pool_data[-1].index)
        all_times = sorted(all_times)

        for pool_data in all_pool_data:
            for time in all_times:
                if time in pool_data.index:
                    nan_filtered_values = [v for v in pool_data.loc[time] if not np.isnan(v)]
                else:
                    nan_filtered_values = []
                data_pooled_per_time.setdefault(time, []).append(nan_filtered_values)

        group_names = [v[0] for v in selected_groups]

        p_values_per_time = collections.OrderedDict()
        for time, values_per_group in data_pooled_per_time.items():
            if len(values_per_group) < 2:
                p_values_per_time[time] = pd.DataFrame(np.nan, index=group_names, columns=group_names)
            else:
                # If one of the group has no valid values, set the Dunn matrix to nan
                if [] in values_per_group:
                    p_values_per_time[time] = pd.DataFrame(np.nan, index=group_names, columns=group_names)
                else:
                    p_values_per_time[time] = pd.DataFrame(sk.posthoc_dunn(values_per_group).to_numpy(), index=group_names, columns=group_names)

        return p_values_per_time

    def evaluate_pairwise_time_effect(self, selected_property):
        """Evaluate the time effect pairwisly for a given property for the selected groups.

        Args:
            selected_property (str): the selected property

        Returns:
            collections.OrderedDict: dict whose key are the name of the selected group and the value the p-values resulting 
            from the Dunn posthocs test stored as a pandas.DataFrame
        """

        selected_groups = [(k, v[0]) for k, v in self._groups.items() if v[1]]
        if len(selected_groups) < 1:
            logging.error('No group selected for evaluating the global time effect')
            return

        p_values = collections.OrderedDict()
        for group, animals_pool_model in selected_groups:
            try:
                dunn_matrix = animals_pool_model.evaluate_pairwise_time_effect(selected_property)
            except InvalidPoolData as error:
                logging.error(str(error))
                continue

            p_values[group] = dunn_matrix

        return p_values

    def export_statistics(self, filename, selected_property='APs'):
        """Export basic statistics (average, median, std, quartile ...) for each group and interval to an excel file.

        Args:
            filename (str): the output excel filename
            selected_property (str): the selected property
            selected_groups (list of str): the selected group. if None all the groups of the model will be used.
        """

        selected_groups = [(k, v[0]) for k, v in self._groups.items() if v[1]]
        if not selected_groups:
            logging.error('No group selected for export')
            return

        workbook = openpyxl.Workbook()
        # Remove the first empty sheet created by default
        workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))

        for group, animals_pool_model in selected_groups:

            try:
                reduced_data = animals_pool_model.get_reduced_data(selected_property)
            except InvalidPoolData:
                logging.error('Could not get reduced pool data for group {}. Skip it.'.format(group))
                continue

            reduced_data.sort_index(inplace=True)

            # Create the excel worksheet
            workbook.create_sheet(group)
            worksheet = workbook.get_sheet_by_name(group)

            worksheet.cell(row=1, column=1).value = 'time'
            worksheet.cell(row=1, column=12).value = 'selected property'
            worksheet.cell(row=2, column=12).value = selected_property
            worksheet.cell(row=1, column=14).value = 'animals'

            # Add titles
            for col, func in enumerate(statistical_functions.keys()):
                worksheet.cell(row=1, column=col+2).value = func

                values = reduced_data[func]
                for row, value in enumerate(values):
                    worksheet.cell(row=row+2, column=1).value = values.index[row]
                    worksheet.cell(row=row+2, column=col+2).value = value

            animals = self._groups[group][0].animals
            for row, animal in enumerate(animals):
                worksheet.cell(row=row+2, column=14).value = animal

        try:
            workbook.save(filename)
        except PermissionError as error:
            logging.error(str(error))
            return

        logging.info('Exported successfully groups statistics in {} file'.format(filename))

    def flags(self, index):
        """Return the flag for specified item.

        Returns:
            int: the flag
        """

        default_flags = super(AnimalsGroupsModel, self).flags(index)

        return QtCore.Qt.ItemIsUserCheckable | default_flags

    def get_animals_pool_model(self, name):
        """Get the animals pool model with a given name.

        Args:
            name (str): the model name

        Returns:
            pigcell.gui.models.animals_pool_model.AnimalsPoolModel: the animals pool model
        """

        if name in self._groups:
            return self._groups[name][0]
        else:
            return None

    def get_data_per_group(self, selected_property):
        """Reduced the pool data for each group for a given property according a given statistics.

        Args:
            selected_property (str): the property
            selected_statistics (str): the statistics

        Returns:
            collections.OrderedDict: the reduced data per group
        """

        data_per_group = collections.OrderedDict()

        groups = list(self._groups.keys())

        for group in groups:
            animals_pool_model, selected = self._groups[group]
            # If the group is selected, compute the reduced stastitics
            if selected:
                data = animals_pool_model.get_pool_data(selected_property)
                data_per_group[group] = data

        return data_per_group

    def get_reduced_data_per_group(self, selected_property, selected_statistics=None):
        """Reduced the pool data for each group for a given property according a given statistics.

        Args:
            selected_property (str): the property
            selected_statistics (str): the statistics

        Returns:
            collections.OrderedDict: the reduced data per group. Stored as a dictionary that map a statistics with the a pandas.DataFrame
        """

        reduced_data_per_group = collections.OrderedDict()

        groups = list(self._groups.keys())

        for group in groups:
            animals_pool_model, selected = self._groups[group]
            # If the group is selected, compute the reduced stastitics
            if selected:
                reduced_data = animals_pool_model.get_reduced_data(selected_property, selected_statistics)
                reduced_data_per_group[group] = reduced_data

        return reduced_data_per_group

    def remove_animals_pool_model(self, model):
        """
        """

        for i, v in enumerate(self._groups.items()):
            name, current_model = v
            if current_model[0] == model:
                self.beginRemoveRows(QtCore.QModelIndex(), i, i)
                del self._groups[name]
                self.endRemoveRows()
                break

    def rowCount(self, parent=None):
        """Returns the number of row of the model.

        Returns:
            int: the number of rows
        """

        return len(self._groups)

    def setData(self, index, value, role):
        """Set the data for a given index and given role.

        Args:
            value (QtCore.QVariant): the data
        """

        if not index.isValid():
            return QtCore.QVariant()

        row = index.row()

        group_names = list(self._groups.keys())

        group = group_names[row]

        if role == QtCore.Qt.CheckStateRole:
            self._groups[group][1] = True if value == QtCore.Qt.Checked else False
            return True

        return super(AnimalsGroupsModel, self).setData(index, value, role)
