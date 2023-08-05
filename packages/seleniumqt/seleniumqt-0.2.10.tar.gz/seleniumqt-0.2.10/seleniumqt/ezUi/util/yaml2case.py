# coding: utf-8
import os
import datetime, time
from seleniumqt.ezUi.util import yaml_data
import openpyxl
from pathlib import Path
import traceback
from seleniumqt.ezUi import run_config

# 根据test_case下文件夹名返回该文件夹下的用例名列表
def case_file(pn):
    business = list()

    try:
        cf = run_config.run_path + '/test_case/' + pn + '/case_edit/'
        if Path(cf).is_dir():
            for i in os.listdir(cf):
                if i.startswith('testCase_') and i.endswith('.xlsx'):
                    business.append(i)
    except:
        traceback.print_exc()

    return business


# 将用例excel内容以列表形式返回，一行对应一个字典
def file2dict(pn):
    case_list = list()

    try:
        cf = case_file(pn)
        for i in cf:
            wb = openpyxl.load_workbook(run_config.run_path + '/test_case/' + pn + '/case_edit/' + i)
            sheets = wb.sheetnames

            for j in sheets:
                ws = wb[j]
                for k in range(ws.max_row - 1):
                    row_dict = dict()
                    row_dict['excel'] = i[9:-5]
                    row_dict['sheet'] = j
                    for l in range(ws.max_column):
                        title = ws.cell(row=1, column=l + 1).value
                        content = ws.cell(row=k + 2, column=l + 1).value
                        row_dict[title] = content
                    case_list.append(row_dict)

    except:
        traceback.print_exc()

    return case_list


def cell2flow(opeStep, pam, expRzlt):
    opeStep_dict = dict()
    pam_dict = dict()
    expRzlt_dict = dict()

    try:

        opeStep_list = opeStep.split('\n')
        pam_list = pam.split('\n')

        for i in opeStep_list:
            if i:
                opeStep_dict[int(i[:i.index('.')]) - 1] = i[i.index('.') + 1:]

        for i in pam_list:
            pam_dict[int(i[:i.index('.')]) - 1] = i[i.index('.') + 1:]

        if expRzlt:
            expRzlt_list = str(expRzlt).split('\n')
            for i in expRzlt_list:
                if i:
                    expRzlt_dict[int(i[:i.index('.')]) - 1] = i[i.index('.') + 1:]

        for i in opeStep_dict.keys():
            if i not in pam_dict.keys():
                pam_dict[i] = ''
        for i in opeStep_dict.keys():
            if i not in expRzlt_dict.keys():
                expRzlt_dict[i] = ''

    except:
        traceback.print_exc()

    return opeStep_dict, pam_dict, expRzlt_dict


def flow2step(opeStep_dict, pam_dict, expRzlt_dict, pn, param_dict):
    step_list = list()

    for i in range(len(opeStep_dict.keys())):
        try:
            # split =pn
            split = 'UAT' if 'uat' in pn else 'QA'

            yaml_pth = run_config.run_path + '/yzf_object/' + split
            yaml_file = opeStep_dict[i].split(':')[0]

            caseObj = yaml_data.get_yaml_data(yaml_pth + '/case_object/' + yaml_file)
            caseStep = yaml_data.get_yaml_data(yaml_pth + '/case_step/' + yaml_file)

            url = caseStep['url']
            pl = opeStep_dict[i].split(':')[0]
            fl = opeStep_dict[i].split(':')[1]
            if fl not in caseStep.keys():
                print('*******KEY not In', fl)
                print('caseStep__KEY', caseStep.keys())
                return
            flow_ = caseStep[fl]
            pam_list_ = pam_dict[i].split(':')

            expRzlt_list_ = expRzlt_dict[i].split(':')
            step_list.append(['url', url, pl, fl])

            def flow2dict(fl):
                try:
                    for fi in fl:
                        if type(fi) is dict:
                            for k, v in fi.items():
                                # try:
                                #     for fj in v:
                                #         print(fi)
                                # except:
                                #     traceback.print_exc()
                                #     print('**==',fl)
                                #     print('**==',fi)

                                for fj in v:
                                    for k_, v_ in fj.items():
                                        ope_obj, param_obj = param2step(v_)
                                        object_value = None
                                        inputValue = param_obj

                                        try:
                                            object_value = caseObj[k]['object'][k_]
                                        except:
                                            print('====', fi)
                                            print('****', fj)
                                            print('*1', caseObj)
                                            print(k, v)
                                            if k in caseObj.keys():
                                                print('*2', caseObj[k])
                                                print(k_, v_)
                                                if k_ in caseObj[k].keys():
                                                    print('*3', caseObj[k]['object'])

                                            os._exit(1)

                                        try:
                                            if 'param_' == inputValue:
                                                inputValue = pam_list_[0]
                                                pam_list_.pop(0)
                                            elif 'exp_' == inputValue:
                                                inputValue = expRzlt_list_[0]
                                                expRzlt_list_.pop(0)
                                        except:

                                            for i_step in step_list:
                                                print(i_step)

                                            time.sleep(1)
                                            print('*' * 8)
                                            traceback.print_exc()

                                            os._exit(1)

                                        step_list.append([object_value, ope_obj, inputValue, k_])

                        elif type(fi) is str:
                            step_list.append([fi, 'add_flow', '*', '*'])
                            fi = caseStep[fi]
                            flow2dict(fi)
                        else:
                            print(fi)
                except:
                    traceback.print_exc()

            flow2dict(flow_)

        except:
            traceback.print_exc()

    # pam = None if 'param_' not in v_ else pam_list_[pam_index]
    # exp = None if 'exp_' not in v_ else expRzlt_list_[exp_index]
    # if pam == '' or exp == '': continue

    step_list_ = list()

    if type(param_dict) is dict:
        for sl in step_list:
            try:
                inputValue_ = sl[2]
                object_value_ = sl[0]

                if inputValue_ and '<' in inputValue_:
                    for k, v in param_dict.items():
                        inputValue_ = inputValue_.replace('<' + k + '>', v)

                if object_value_ and '<' in object_value_:
                    for k, v in param_dict.items():
                        object_value_ = object_value_.replace('<' + k + '>', v)

                step_list_.append([object_value_, sl[1], inputValue_, sl[3]])
            except:
                traceback.print_exc()
    return step_list_


def param2step(vl):
    try:
        param_obj = None

        if 'exp' in vl:
            ope_obj = '验证'
            if str(vl).startswith('exp=='):
                param_obj = vl[5:]
            elif 'exp_' in vl:
                param_obj = 'exp_'
        elif 'param_' in vl:
            ope_obj = vl.replace('param_', '')
            param_obj = 'param_'
        elif str(vl).startswith('<') and str(vl).endswith('>'):
            ope_obj = '取参'
            param_obj = vl[1:-1]
        elif '输入' in vl:
            ope_obj = '输入'
            param_obj = vl[2:]
        elif '选择' in vl:
            ope_obj = '选择'
            param_obj = vl[2:]
        elif '上传文件' in vl:
            ope_obj = '上传文件'
            param_obj = vl[4:]
        else:
            ope_obj = vl

        return ope_obj, param_obj
    except:
        traceback.print_exc()
        os._exit(1)


def run(gid, param_dict):
    user_case = list()
    try:

        for pn in os.listdir(run_config.run_path + '/test_case'):
            # 与jenkins部署的job名有关，用以控制jenkins执行哪个用例
            # 如果test_case下文件夹名与jenkins的job名一致则继续执行，不一致则跳过本次循环

            if run_config.run_case:
                if pn not in run_config.run_case:
                    continue
            elif not run_config.run_path.endswith(pn):
                continue

            case_list = file2dict(pn)

            for i in case_list:
                if int(i['优先级']) not in gid:
                    continue

                case_ = dict()
                case_['business'] = i['excel']
                case_['module'] = i['sheet']
                case_['id'] = i['序号']
                case_['name'] = i['概述']
                case_['systerm'] = pn
                case_['requirement'] = i['需求']
                case_['priority'] = i['优先级']

                o, p, e = cell2flow(i['操作步骤'], i['参数'], i['预期结果'])
                step_list = list()

                t_now = datetime.datetime.now()
                t_today = time.strftime("%Y-%m-%d", time.localtime())

                flow2step_ = flow2step(o, p, e, pn, param_dict)

                # print('flow2step_', flow2step_)
                for isl in flow2step_:
                    try:
                        kv = list()
                        for step_ in isl:
                            if type(step_) is str and '_scd' in step_:
                                step_ = str(step_).replace('_scd',
                                                           str(t_now.year)
                                                           + str(t_now.month)
                                                           + str(t_now.day)
                                                           + str(t_now.hour)
                                                           + str(t_now.minute)
                                                           + 'yzf')
                            if type(step_) is str and '_time10' in step_:
                                step_ = str(step_).replace('_time10', str(int(int(t_now.timestamp()))))
                                # step_ = str(step_).replace('_time10', '001')

                            if type(step_) is str and '_today' in step_:
                                step_ = str(step_).replace('_today', str(t_today))

                            kv.append(step_)

                        if kv[2] == 'NA' or kv[2] == '':
                            continue

                        if '_NN_' in case_['id'] and kv[1] == '验证' and 'wait_Exit' != kv[2]:
                            print('****', kv, )
                        else:
                            step_list.append(kv)
                            # if 'url' == kv[0]:
                            #     print('\n' * 3)
                            # if 'add_flow' == kv[1]:
                            #     print('*' * 88)
                            # print('----', len(step_list), kv)
                    except:
                        traceback.print_exc()

                case_['Step'] = step_list
                user_case.append(case_)

    except:
        traceback.print_exc()

    return user_case


if __name__ == '__main__':
    print("*****")
    ls = case_file("MerchantAccess")
    print(ls)
    print("##############")
    aa = file2dict("MerchantAccess")
    print(type(aa))
    print(aa)
