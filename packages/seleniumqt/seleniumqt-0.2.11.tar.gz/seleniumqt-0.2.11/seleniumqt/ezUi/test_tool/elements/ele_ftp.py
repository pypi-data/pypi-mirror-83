# coding: utf-8
import os
import time
import logging
import traceback
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def elements(p):
    try:
        text_value = p[1][0]
        if str(text_value).startswith('FTP上传文件'):
            oid = int(round(time.time() * 1000))

            excel_up = '%s_20200707.xlsx' % (oid)
            dir_path = os.path.dirname(os.path.abspath(__file__)) + '/../../conf/'

            excel_up_path = dir_path + 'tmp/' + excel_up
            if str(text_value).endswith('_'):
                fn_ = 'data_file/ftp_demo_.xlsx'
            else:
                fn_ = 'data_file/ftp_demo.xlsx'
            wb = load_workbook(filename=dir_path + fn_)

            ws = wb['批量转账模板']
            ws["A2"] = str(oid - 1)
            ws['A3'] = str(oid - 2)

            wb.save(excel_up_path)

            pkey = dir_path + 'data_file/898110311.0'

            import paramiko

            key = paramiko.RSAKey.from_private_key_file(pkey, password='898110311')
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

            ssh.connect('172.23.248.4', username='898110311', password='898110311', pkey=key)

            t = ssh.get_transport()
            sftp = paramiko.SFTPClient.from_transport(t)
            sftp.put(excel_up_path, "/transfer/domestic/details/" + excel_up)

            t.close()

            return str(oid)
    except:
        traceback.print_exc()
