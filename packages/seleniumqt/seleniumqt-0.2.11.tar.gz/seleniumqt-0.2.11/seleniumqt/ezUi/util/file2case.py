# coding: utf-8
import os
import datetime, time
from seleniumqt.ezUi.util import yaml_data
import openpyxl
from pathlib import Path


# 获取并返回项目路径
def pt():
    p = os.getcwd()

    try:
        if 'interfacetest' in p:
            p = p[:p.index('interfacetest') + 13]
    except Exception as e:
        print(e)
        raise Exception("函数::pt")

    return p


# 根据test_case下文件夹名返回该文件夹下的用例名列表
def case_file(pn):
    business = list()

    try:
        cf = pt() + '/test_case/' + pn + '/case_edit/'
        if Path(cf).is_dir():
            for i in os.listdir(cf):
                if i.startswith('testCase_') and i.endswith('.xlsx'):
                    business.append(i)
    except Exception as e:
        print(e)
        raise Exception("函数::case_file")

    return business


# 将用例excel内容以列表形式返回，一行对应一个字典
def file2dict(pn):
    case_list = list()

    try:
        cf = case_file(pn)
        for i in cf:
            wb = openpyxl.load_workbook(pt() + '/test_case/' + pn + '/case_edit/' + i)
            sheets = wb.sheetnames

            for j in sheets:
                ws = wb[j]
                for k in range(ws.max_row - 1):
                    row_dict = dict()
                    row_dict['excel'] = i[9:-5]
                    row_dict['sheet'] = j
                    for l in range(ws.max_column):
                        title = ws.cell(row=1, column=l + 1).value
                        content = ws.cell(row=k + 2, column=l + 1).value
                        row_dict[title] = content
                    case_list.append(row_dict)

    except Exception as e:
        print(e)
        raise Exception("函数::file2dict")

    return case_list


def cell2flow(opeStep, pam, expRzlt):
    opeStep_dict = dict()
    pam_dict = dict()
    expRzlt_dict = dict()

    try:

        opeStep_list = opeStep.split('\n')
        pam_list = pam.split('\n')

        for i in opeStep_list:
            if i:
                opeStep_dict[int(i[:i.index('.')]) - 1] = i[i.index('.') + 1:]

        for i in pam_list:
            pam_dict[int(i[:i.index('.')]) - 1] = i[i.index('.') + 1:]

        if expRzlt:
            expRzlt_list = str(expRzlt).split('\n')
            for i in expRzlt_list:
                if i:
                    expRzlt_dict[int(i[:i.index('.')]) - 1] = i[i.index('.') + 1:]

        for i in opeStep_dict.keys():
            if i not in pam_dict.keys():
                pam_dict[i] = ''
        for i in opeStep_dict.keys():
            if i not in expRzlt_dict.keys():
                expRzlt_dict[i] = ''

    except Exception as e:
        print(e)
        raise Exception("函数::cell2flow")

    return opeStep_dict, pam_dict, expRzlt_dict


def flow2step(opeStep_dict, pam_dict, expRzlt_dict, pn, param_dict):
    step_list = list()
    for i in range(len(opeStep_dict.keys())):
        try:
            # caseObj = yaml_data.get_yaml_data(
            #     pt() + '/test_case/' + pn + '/case_object/' + opeStep_dict[i].split(':')[0])
            # caseStep = yaml_data.get_yaml_data(
            #     pt() + '/test_case/' + pn + '/case_step/' + opeStep_dict[i].split(':')[0])

            caseObj = yaml_data.get_yaml_data(
                pt() + '/yzf_object/case_object/' + opeStep_dict[i].split(':')[0])
            caseStep = yaml_data.get_yaml_data(
                pt() + '/yzf_object/case_step/' + opeStep_dict[i].split(':')[0])

            url = caseStep['url']
            pl = opeStep_dict[i].split(':')[0]
            fl = opeStep_dict[i].split(':')[1]
            flow_ = caseStep[fl]
            pam_list_ = pam_dict[i].split(':')
            expRzlt_list_ = expRzlt_dict[i].split(':')
            step_list.append(['url', url, pl, fl])

            def flow2dict(fl):
                try:
                    for fi in fl:

                        if type(fi) is dict:
                            for k, v in fi.items():
                                for fj in v:
                                    for k_, v_ in fj.items():
                                        if not v_:
                                            print('*' * 8, k_, v_)
                                        if k_ == 'skip':
                                            step_list.append(['skip', '跳过', None, v_])
                                            continue

                                        if 'param' in v_ and len(pam_list_) < int(v_.split('param')[1]) - 1:
                                            # print('*' * 8)
                                            # print(v_)
                                            # print(pam_list_)
                                            # print(int(v_.split('param')[1]) - 1)
                                            # print('-' * 8)
                                            break

                                        if 'param' in v_:

                                            # print('**1', pam_list_)
                                            # print('**2', v_.split('param'))
                                            # print('**3', int(v_.split('param')[1]) - 1)
                                            # print('**4', pam_list_[int(v_.split('param')[1]) - 1])

                                            if pam_list_[int(v_.split('param')[1]) - 1] == '_':
                                                continue

                                        ope_obj, param_obj = param2step(v_, pam_list_, expRzlt_list_)

                                        inputValue = param_obj
                                        object_value = None
                                        try:
                                            object_value = caseObj[k]['object'][k_]

                                        except:
                                            print('*1', ope_obj)
                                            print('*2', param_obj)
                                            print('*3', k)
                                            print('*4', k_)
                                            print('*5', caseObj)
                                            print('*6', caseObj[k])
                                            print('*7', caseObj[k]['object'])
                                            print('*8', caseObj[k]['object'][k_])

                                        step_list.append([object_value, ope_obj, inputValue, k_])

                        elif type(fi) is str:

                            step_list.append([fi, 'add_flow', '*', '*'])

                            fi = caseStep[fi]
                            flow2dict(fi)
                        else:
                            print(fi)
                except Exception as e1:
                    print(e1)
                    print(fl)
                    raise Exception("函数::flow2step")

            flow2dict(flow_)
        except Exception as e:
            print(e)
            print(i, opeStep_dict)

            raise Exception("函数::flow2step")

    step_list_ = list()

    if type(param_dict) is dict:
        for sl in step_list:
            try:
                inputValue_ = sl[2]
                object_value_ = sl[0]

                if inputValue_ and '<' in inputValue_:
                    for k, v in param_dict.items():
                        inputValue_ = inputValue_.replace('<' + k + '>', v)

                if '<' in object_value_:
                    for k, v in param_dict.items():
                        object_value_ = object_value_.replace('<' + k + '>', v)

                step_list_.append([object_value_, sl[1], inputValue_, sl[3]])
            except Exception as e:
                print('Exception', sl)
                raise e
    return step_list_


def param2step(vl, pam_list, expRzlt_list):
    try:
        param_obj = None
        if 'param' in vl:
            ope_obj, pid_obj = vl.split('param')
            param_obj = pam_list[int(pid_obj) - 1]
        elif 'exp' == vl[:3]:
            ope_obj = '验证'
            if str(vl).startswith('exp=='):
                param_obj = vl[5:]
            else:
                pid_rzlt = vl.split('exp')[1]
                param_obj = expRzlt_list[int(pid_rzlt) - 1]

        elif str(vl).startswith('<') and str(vl).endswith('>'):
            ope_obj = '取参'
            param_obj = vl[1:-1]
        elif str(vl).startswith('输入'):
            ope_obj = '输入'
            param_obj = vl[2:]
        elif str(vl).startswith('选择'):
            ope_obj = '选择'
            param_obj = vl[2:]
        else:
            ope_obj = vl

        return ope_obj, param_obj
    except Exception as e:
        print('Exception', e)
        print('步骤', vl)
        print('参数', pam_list)
        print('预期', len(expRzlt_list))
        raise Exception("函数::param2step")


def run(gid, param_dict):
    user_case = list()
    try:

        for pn in os.listdir(pt() + '/test_case'):
            # 与jenkins部署的job名有关，用以控制jenkins执行哪个用例
            # 如果test_case下文件夹名与jenkins的job名一致则继续执行，不一致则跳过本次循环
            # if not pt().endswith(pn):
            #     continue

            # if not pn == 'onlinepay_B2B+B2C':
            #     continue

            # if 'MerchantAccess' != pn:
            #     continue

            # if 'ehking-ert' != pn:
            #     continue

            if 'ehking-agreement' != pn:
                continue

            # if 'ehking-b2b' != pn:
            #     continue

            # if 'ehking-b2c' != pn:
            #     continue

            # if 'ehking-scanCodeUnion' != pn:
            #     continue
            #
            # if 'ehking-scanCodeAlipay' != pn:
            #     continue

            # if 'ehking-transfer' != pn:
            #     continue

            # pn：test_case下的文件夹
            case_list = file2dict(pn)

            for i in case_list:
                # print('*' * 8, gid, int(i['优先级']))
                if int(i['优先级']) not in gid:
                    continue

                case_ = dict()
                case_['business'] = i['excel']
                case_['module'] = i['sheet']
                case_['id'] = i['序号']
                case_['name'] = i['概述']
                case_['systerm'] = pn
                case_['requirement'] = i['需求']
                case_['priority'] = i['优先级']

                o, p, e = cell2flow(i['操作步骤'], i['参数'], i['预期结果'])
                step_list = list()

                t_now = datetime.datetime.now()
                t_today = time.strftime("%Y-%m-%d", time.localtime())

                for isl in flow2step(o, p, e, pn, param_dict):
                    try:
                        kv = list()
                        for step_ in isl:
                            if type(step_) is str and '_scd' in step_:
                                step_ = str(step_).replace('_scd',
                                                           str(t_now.year)
                                                           + str(t_now.month)
                                                           + str(t_now.day)
                                                           + str(t_now.hour)
                                                           + str(t_now.minute)
                                                           + 'yzf')
                            if type(step_) is str and '_time10' in step_:
                                step_ = str(step_).replace('_time10', str(int(int(t_now.timestamp()))))
                                # step_ = str(step_).replace('_time10', '001')

                            if type(step_) is str and '_today' in step_:
                                step_ = str(step_).replace('_today', str(t_today))

                            kv.append(step_)

                        if kv[2] == 'NA':
                            continue

                        if '_NN_' in case_['id'] and kv[1] == '验证' and 'wait_Exit' != kv[2]:
                            print('****', kv, )
                        else:
                            step_list.append(kv)
                            print('----', kv, )
                    except Exception as e:
                        print("isl", isl)
                        print(e)
                        raise Exception("函数::run")
                case_['Step'] = step_list
                user_case.append(case_)

    except Exception as e:
        print(param_dict)
        print(e)
        raise Exception("函数::run")

    return user_case


if __name__ == '__main__':
    print("*****")
    ls = case_file("MerchantAccess")
    print(ls)
    print("##############")
    aa = file2dict("MerchantAccess")
    print(type(aa))
    print(aa)
    # ks = run([1], None)
    # print(len(ks), ks)
    # if ks:
    #     for i in ks:
    #         for k, v in i.items():
    #             if type(v) is list:
    #                 for j in v:
    #                     print(j)
    #             else:
    #                 print(k, v)
