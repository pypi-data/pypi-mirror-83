"""
Integration tests for ELDAM
"""

import os
from datetime import datetime

import openpyxl

from eldam.core.parameters import EldaTemplateParameters
from eldam.core.elda import Elda
from eldam.core.simapro import SimaProXlsxReader, SimaProCsvWriter
from eldam.core.elda_index import EldaIndexer
from eldam.utils.xls import compare_workbook
from eldam.utils.misc import compare_file

from tests.testing_parameters import *


class TestEldaFromXlsx:

    def setup_class(self):
        """ Creates and saves a elda file from the simapro export """
        self.export_file = os.path.join(DATA_FOLDER, 'temporary_file.xlsm')

    def teardown_class(self):
        """ Deletes the elda file """
        os.remove(self.export_file)

    def elda_from_xlsx(self, xlsx, elda):
        """ Asserts the created elda file is the same than the reference """

        xlsx_reader = SimaProXlsxReader(xlsx)
        process = xlsx_reader.to_processes()[0]
        process.remove_unused_parameters()
        generated_elda = Elda()
        generated_elda.add_version(process)
        generated_elda.save(self.export_file)

        output_workbook = openpyxl.load_workbook(self.export_file)
        reference_workbook = openpyxl.load_workbook(elda)

        differing_cells = compare_workbook(output_workbook, reference_workbook)

        assert differing_cells is None or (len(differing_cells) == 1 and differing_cells[0][0][0:2] ==
                                           EldaTemplateParameters().VERSION_DATE_CELL)

    def test_elda_from_xlsx_a(self):
        self.elda_from_xlsx(XLSX_A, ELDA_A)

    def test_elda_from_xlsx_b(self):
        self.elda_from_xlsx(XLSX_B, ELDA_B)

    def test_elda_from_xlsx_d(self):
        self.elda_from_xlsx(XLSX_D, ELDA_D)

    def test_elda_from_xlsx_e(self):
        self.elda_from_xlsx(XLSX_E, ELDA_E)

    def test_elda_from_xlsx_a_with_excel_link(self):
        self.elda_from_xlsx(XLSX_A_WITH_EXCEL_LINK, ELDA_A_WITH_EXCEL_LINK)


class TestCsvFromElda:

    def setup_class(self):
        self.export_file = os.path.join(DATA_FOLDER, 'temporary_file.csv')

    def teardown_class(self):
        os.remove(self.export_file)

    def csv_from_elda(self, ref_elda, ref_csv):
        elda = Elda(ref_elda)
        process_from_file = elda.read_last_version().to_process()
        SimaProCsvWriter(process_from_file).to_csv(self.export_file)

        assert compare_file(self.export_file, ref_csv) == ''

    def test_csv_from_elda_a(self):
        self.csv_from_elda(ELDA_A, CSV_A)

    def test_csv_from_elda_b(self):
        self.csv_from_elda(ELDA_B, CSV_B)

    def test_csv_from_elda_c(self):
        self.csv_from_elda(ELDA_C_MAJOR, CSV_C)

    def test_csv_from_elda_d(self):
        self.csv_from_elda(ELDA_D, CSV_D)

    def test_csv_from_elda_e(self):
        self.csv_from_elda(ELDA_E, CSV_E)


class TestEldaIndex:

    def setup_class(self):
        self.export_file = os.path.join(DATA_FOLDER, 'temporary_file.xlsx')

    def teardown_class(self):
        os.remove(self.export_file)

    def test_elda_index(self):
        indexer = EldaIndexer(os.path.join(DATA_FOLDER, 'elda_index'))
        indexer.build_index()
        indexer.save_index(self.export_file)

        output_workbook = openpyxl.load_workbook(self.export_file)
        reference_workbook = openpyxl.load_workbook(os.path.join(DATA_FOLDER, 'elda_index', 'EldaIndex.xlsx'))

        assert compare_workbook(output_workbook, reference_workbook) is None
