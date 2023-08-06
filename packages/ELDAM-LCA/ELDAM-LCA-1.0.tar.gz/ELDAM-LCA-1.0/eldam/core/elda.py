""" Classes used for Elda handling """

import os
from datetime import datetime
from collections import OrderedDict

import openpyxl
from openpyxl.utils import column_index_from_string, coordinate_to_tuple, get_column_letter, absolute_coordinate
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

from eldam.core.lci_data import *
from eldam.core.parameters import *
from eldam.utils.elda import version_from_text, add_major_version_from_text, add_minor_version_from_text
from eldam.utils.xls import copy_border_style, add_named_cell, make_excel_formula, get_original_formula, \
    copy_cell_style, reset_cells, delete_cells, is_formula_using_cells_adresses
from eldam.utils.styles import *
from eldam.utils.exceptions import NotAnEldaError, EldaVersionError, ExcelFormulaError
from eldam.utils.misc import n_str
from eldam.settings import ELDA_TEMPLATE_VERSION


class Elda:
    """
    Class describing an ELSA LCI formatted elda
    """

    def __init__(self, filepath=None):
        """
        Args:
            filepath (str): Filepath of the elda to update. If None, elda template will be used.
        """

        # Loading current template parameters
        self.elda_template_parameters = EldaTemplateParameters()

        # If no elda has been given, uses the elda template
        if filepath is None:
            filepath = ELDA_TEMPLATE_FILEPATH
            self.is_empty = True
        else:
            self.is_empty = False
            if os.path.splitext(filepath)[1].lower() != '.xlsm':
                raise NotAnEldaError('The file seems not to be an Elda')

        # Opening the elda file or the template
        self.workbook = openpyxl.load_workbook(filename=filepath, keep_vba=True)

        # Checks that the file is an Elda
        if self.workbook.sheetnames[0:3] != ['Status', 'Simapro units', 'V1.0']:
            raise NotAnEldaError('The file seems not to be an Elda')

        # Gets the elda template version (to check compatibility)
        self.template_version = str(self.workbook['Status']
                                    [ELDA_TEMPLATE_VERSION_CELL].value)

        if self.template_version != ELDA_TEMPLATE_VERSION:
            self.update_elda_template()

        # Fix comment size
        self.fix_comment_size()

        # Fix SimaPro units validation
        self.fix_unit_validation()

    @property
    def last_version_number(self):
        """
        Returns the last version number.

        Returns:
            str : Last version number
        """
        return self.workbook.sheetnames[-1]

    def update_elda_template(self):
        """
        Updates an ELDA with the current ELDA template.
        If the ELDA is already based on the current ELDA template, does nothing.
        """

        if self.template_version == ELDA_TEMPLATE_VERSION:
            return

        # Getting all versions of the process
        process_versions = self.read_all_versions_to_processes()

        # Resetting the ELDA with the current ELDA template
        self.__init__()

        # Inserting processes in the new ELDA
        for version in process_versions:
            is_major_version = version.split('.')[1] == '0'
            self.add_version(process=process_versions[version], major_version=is_major_version)

    def add_version(self, process, major_version=False):
        """
        Adds a new version to the elda

        Args:
            process (Process): Process to add to the new version
            major_version (bool): Is it a major version (2.0, 3.0, ...) or a minor version (2.1, 2.2, 2.3, ...)
        """
        # Calculating version number
        if self.is_empty:
            new_version_number = "V1.0"

        else:
            # The last version is always the last item of self.workbook.sheetnames
            previous_version_number = self.workbook.sheetnames[-1]
            previous_version = self.workbook[previous_version_number]

            if major_version:
                new_version_number = add_major_version_from_text(previous_version_number)
            else:
                new_version_number = add_minor_version_from_text(previous_version_number)

            new_version = self.workbook.copy_worksheet(previous_version)

            # Renaming the copyied version
            new_version.title = new_version_number

            # Fixing conditional formatting loss
            new_version.conditional_formatting = previous_version.conditional_formatting

            # Fixing data validation loss
            new_version.data_validations = previous_version.data_validations

            # Fixing grid style loss
            new_version.sheet_view.showGridLines = previous_version.sheet_view.showGridLines

            # Fixing merged cells style on previous version
            self.fix_style()

        version_writer = EldaVersionWriter(elda=self, version=new_version_number)
        version_writer.add_process_data(process)
        self.is_empty = False

    def update_last_version(self, process):
        """
        Updates last version with a new Process

        Args:
            process (Process): Process to use for the update
        """

        # Creating a EldaVersionWriter on this sheet and adding data
        version_writer = EldaVersionWriter(elda=self, version=self.last_version_number)
        version_writer.add_process_data(process)

        self.fix_style()

    def update_last_version_metadata(self, process, attributes=ATTRIBUTES_TO_COPY):
        """
        Updates only the metadata of the Elda's last version (except the process name and synonym).
        This is useful to copy metadata from an Elda to other Eldas of the same dataset.

        Args:
            process (Process): Process to copy metadata from
            attributes (list): List of process attributes names to copy
        """

        # Creating a EldaVersionWriter on this sheet and adding metadata
        version_writer = EldaVersionWriter(elda=self, version=self.last_version_number)
        version_writer.add_process_metadata(process, attributes)

        self.fix_style()

    def read_version(self, version):
        """
        Reads specified version

        Args:
            version (str): Version name

        Returns:
            EldaVersionReader: Version reader
        """
        if version not in self.workbook.sheetnames:
            raise EldaVersionError(f"ELDA has no version with number {version}")

        return EldaVersionReader(self, version)

    def read_last_version(self):
        """
        Reads last version of the Elda in a EldaVersionReader object

        Returns:
            EldaVersionReader: Last version
        """

        # Reading this sheet
        return self.read_version(self.last_version_number)

    def read_all_versions_to_processes(self):
        """
        Reads all the versions of a ELDA and returns an ordered dict containing versions numbers as keys
        and version processes as values

        Returns:
            OrderedDict:  Ordered dict containing versions numbers as keys and version processes as values
        """
        result = OrderedDict()

        for version in self.workbook.sheetnames[2:]:
            result[version] = self.read_version(version).to_process()

        return result

    def save(self, filepath):
        """
        Exports the elda to a file

        Args:
            filepath (str): Filepath to export the elda
        """
        # Sets the active worksheet to the status version
        self.workbook.active = 0

        # Deselecting other sheets
        for sheet in self.workbook.worksheets:
            sheet.sheet_view.tabSelected = False

        # Reenable sheet protection
        for sheet in self.workbook.worksheets:
            sheet.protection.formatCells = False
            sheet.protection.formatRows = False
            sheet.protection.formatColumns = False
            sheet.protection.insertHyperlinks = False
            sheet.protection.enable()

        self.workbook.save(filepath)

    def fix_style(self):
        """ Fixes style issues """
        # For all sheets of the Elda and for all cells in sheet.merged_cells, copies the border style of the first cell
        # to the others
        for sheet in self.workbook.worksheets[2:]:
            for merged_cells in sheet.merged_cells:
                reference_cell = str(merged_cells).split(':')[0]
                copy_border_style(sheet, reference_cell, str(merged_cells))

            # Setting the priority of the changed values conditional formatting to 1
            [x for x in sheet.conditional_formatting._cf_rules.values()
             if x[0].formula[0] == self.elda_template_parameters.CHANGED_VALUE_CF_FORMULA][0][0].priority = 1

    def fix_comment_size(self):
        """
            Fixes cell comments sizes

            On reading a workbook, openpyxl looses comments positions and size. They must be reset manually.
        """
        for sheet in self.workbook.worksheets[2:]:
            for cell in self.elda_template_parameters.ELDA_COMMENTS:
                comment = sheet[cell].comment

                if comment is not None:
                    comment.width, comment.height = self.elda_template_parameters.COMMENT_DIMENSIONS[
                        self.elda_template_parameters.ELDA_COMMENTS[cell]]

    def fix_unit_validation(self):
        """
            Fixes units data validation

            On reading a workbook, openpyxl looses the unit validation rule. This method reset it.
        """

        dv = DataValidation(type="list",
                            formula1=self.elda_template_parameters.SIMAPRO_UNITS_DEFINITION_CELLS)
        dv.add("{}{}:{}{}".format(self.elda_template_parameters.FLOW_FIELDS_COLUMNS['unit'],
                                  self.elda_template_parameters.FIRST_FLOW_ROW_NUMBER,
                                  self.elda_template_parameters.FLOW_FIELDS_COLUMNS['unit'],
                                  self.elda_template_parameters.LAST_FLOW_ROW_NUMBER))

        self.workbook['V1.0'].add_data_validation(dv)

class EldaVersionWriter:
    """ Class used to create an ELSA LCI formatted sheet (a version of the elda). """

    def __init__(self, elda, version):
        """
        Args:
            elda (Elda): Elda to which the SimaPro export should be added
            version (str): Version number
        """

        # Loading current template parameters
        self.elda_template_parameters = EldaTemplateParameters()

        self.major, self.minor = version_from_text(version)
        self.elda = elda
        self.version = version
        self.last_input_parameter_cell_coords = self.elda_template_parameters.DEFAULT_LAST_INPUT_PARAMETER_CELL_COORDS
        self.last_calculated_parameter_cell_coords = \
            self.elda_template_parameters.DEFAULT_LAST_CALCULATED_PARAMETER_CELL_COORDS
        self.inserted_input_parameters = 0
        self.inserted_calculated_parameters = 0
        self.available_input_parameter_rows = self.elda_template_parameters.INPUT_PARAMETERS_BLOCK_HEIGHT - 1
        self.available_calculated_parameter_rows = self.elda_template_parameters.CALCULATED_PARAMETERS_BLOCK_HEIGHT - 1

        # Getting the corresponding sheet
        self.sheet = self.elda.workbook[version]

    def _add_attribute(self, process, attribute):
        """
        Adds a process attribute data and metadata. Used by add_process_metadata() and add_process_data()

        Args:
            process (Process):
            attribute (tuple):
        """
        # Inserting attribute value
        self.sheet[attribute[1]] = getattr(process, attribute[0])

        # Inserting attribute review data
        attribute_dict = process.metadata_review.get(attribute[0], {})
        if attribute[1] != self.elda_template_parameters.GENERAL_COMMENT_CELL:
            self.sheet[attribute[1].replace('D', 'E')] = attribute_dict.get('comment', None)
        self.sheet[attribute[1].replace('D', 'F')] = attribute_dict.get('comment_for_reviewer', None)
        self.sheet[attribute[1].replace('D', 'G')] = attribute_dict.get('review_state', 0) or 0
        self.sheet[attribute[1].replace('D', 'I')] = attribute_dict.get('reviewer_comment', None)

    def add_process_metadata(self, process, attributes=ATTRIBUTES_TO_COPY):
        """
        Adds or overwrite the process metadata only to the Elda.

        Args:
            process (Process): Process of which the metadata are added
            attributes (list): List of process attributes names to copy
        """
        # METADATA
        self.sheet[self.elda_template_parameters.MAJOR_VERSION_NUMBER_CELL] = self.major
        self.sheet[self.elda_template_parameters.MINOR_VERSION_NUMBER_CELL] = self.minor
        self.sheet[self.elda_template_parameters.NEXT_MINOR_VERSION_NUMBER_CELL] = self.minor + 1

        self.sheet[self.elda_template_parameters.VERSION_DATE_CELL] = process.date or datetime.now().strftime(
            "%d/%m/%Y")

        # Looping on attributes to insert them
        for attribute in self.elda_template_parameters.METADATA_CELLS.items():

            # Skipping date as it has already been inserted
            # Skipping name and synonym as they must be inserted by add_process_data()
            if attribute[0] not in attributes:
                continue

            self._add_attribute(process=process, attribute=attribute)

        # Looping on version info cells to insert them
        for attribute in self.elda_template_parameters.VERSION_INFO_CELLS.items():
            # Inserting attribute value
            self.sheet[attribute[1]] = getattr(process, attribute[0])

    def add_process_data(self, process):
        """
        Adds or overwrite the process data and metadata to the Elda

        Args:
            process (Process): Process to add to the Elda
        """

        # Adding metadata
        self.add_process_metadata(process)

        # Adding input and output masses
        attributes = [(k, v) for (k, v) in self.elda_template_parameters.METADATA_CELLS.items()
                      if k in ('input_mass', 'output_mass')]
        for attribute in attributes:
            self._add_attribute(process=process, attribute=attribute)

        # PARAMETERS
        # Removing parameters from the sheet
        # This step is necessary because otherwise, self.insert_parameter() will insert after the existing parameters
        # instead of overwriting them.

        self.clean_data()
        for parameter in process.parameters:
            self.insert_parameter(parameter)

        # FLOWS
        # Looping on each flow category to load the data into the elda
        flow_row_counter = self.elda_template_parameters.FIRST_FLOW_ROW_NUMBER

        for flow in process.flows:
            # Inserting data in the right cell
            for name, value in flow.__dict__.items():

                if name in self.elda_template_parameters.FLOW_FIELDS_COLUMNS:
                    coord = self.elda_template_parameters.FLOW_FIELDS_COLUMNS[name] + str(flow_row_counter)

                    # Converting formula
                    if (name == 'amount' and flow.amount_type == 'Formula') or (name == 'allocation'):
                        self.sheet[coord] = make_excel_formula(value)
                    else:
                        self.sheet[coord] = value

            flow_row_counter += 1  # Incrementing the flow row counter by 1

        # FIXING MERGED CELLS STYLE
        self.elda.fix_style()

        # Manually fixing a merge style problem
        self.sheet[self.elda_template_parameters.REVIEW_STATE_CELL].border.right.style = \
            self.sheet[self.elda_template_parameters.REVIEW_STATE_CELL].border.top.style

    def add_parameter_block(self, parameter_type):
        """
        Adds a new input or calculated parameter block

        Args:
            parameter_type (str): 'Input' or 'Calculated'
        """
        parameter_type = parameter_type.lower()
        assert parameter_type in ('input', 'calculated')

        if parameter_type == 'input':
            # Updating the number of available input parameter rows
            self.available_input_parameter_rows += self.elda_template_parameters.INPUT_PARAMETERS_BLOCK_HEIGHT - 1
            block_height = self.elda_template_parameters.INPUT_PARAMETERS_BLOCK_HEIGHT
        else:
            self.available_calculated_parameter_rows += \
                self.elda_template_parameters.CALCULATED_PARAMETERS_BLOCK_HEIGHT - 1
            block_height = self.elda_template_parameters.CALCULATED_PARAMETERS_BLOCK_HEIGHT

        # As it is larger than calc parameters block width it is the default for both
        block_width = self.elda_template_parameters.INPUT_PARAMETERS_BLOCK_WIDTH

        # Copying the last parameter block
        # Getting cells to copy
        last_param_row, last_param_col = coordinate_to_tuple(self.last_input_parameter_cell_coords
                                                             if parameter_type == 'input'
                                                             else self.last_calculated_parameter_cell_coords)

        min_col = last_param_col
        max_col = last_param_col + block_width - 1
        min_row = last_param_row - block_height + 1
        max_row = last_param_row

        cells_to_copy = list(self.sheet.iter_rows(min_row=min_row,
                                                  max_row=max_row,
                                                  min_col=min_col,
                                                  max_col=max_col))

        # Copying the cells
        for row in cells_to_copy:
            # Ignoring merged cells
            for cell in row:
                if type(cell) == openpyxl.cell.cell.MergedCell:
                    continue
                destination_cell = self.sheet.cell(row=cell.row,
                                                   column=cell.col_idx + block_width + 1)
                copy_cell_style(reference_cell=cell, destination_cell=destination_cell)

                if cell.row == min_row:  # Headers
                    destination_cell.value = cell.value
                elif cell.col_idx == min_col + 6:  # Review column
                    destination_cell.value = '=IF({minor_version_number}=0,"",IF({col}{row}=0,' \
                                             'IF({namecol}{row}<>"","Invalid",""),' \
                                             'IF({col}{row}=1,"Needs changes",IF({col}{row}=2,"Valid",""))))'. \
                        format(minor_version_number=self.elda_template_parameters.MINOR_VERSION_NUMBER_FIXED_CELL,
                               namecol=get_column_letter(min_col + block_width + 1),
                               col=get_column_letter(min_col + 5 + block_width + 1),
                               row=cell.row)
                elif parameter_type == 'calculated' and cell.col_idx == min_col + 2:  # Fomula column
                    destination_cell.value = '=IF({namecol}{row}="","",IF(_xlfn.ISFORMULA({col}{row}),' \
                                             'RIGHT(_xlfn.FORMULATEXT({col}{row}),' \
                                             'LEN(_xlfn.FORMULATEXT({col}{row}))-1),{col}{row}))'. \
                        format(namecol=get_column_letter(min_col + block_width + 1),
                               col=get_column_letter(min_col + 1 + block_width + 1),
                               row=cell.row)
                elif cell.col_idx == min_col + 5:
                    destination_cell.value = 0

        # Merging review header
        self.sheet.merge_cells(start_row=min_row,
                               end_row=min_row,
                               start_column=min_col + 5 + block_width + 1,
                               end_column=min_col + 6 + block_width + 1)

        if parameter_type == 'input':
            # Merging value column
            for row_num in range(0, block_height):
                self.sheet.merge_cells(start_row=min_row + row_num,
                                       end_row=min_row + row_num,
                                       start_column=min_col + 1 + block_width + 1,
                                       end_column=min_col + 2 + block_width + 1)

        # Adding conditional formatting
        # - Changed values
        # cell_range = "${min_col1}${min_row}:${max_col1}${max_row}".format(
        cell_range = "${min_col1}${min_row}:${max_col1}${max_row} ${min_col2}${min_row}:${max_col2}${max_row}".format(
            min_col1=get_column_letter(min_col + block_width + 1),
            min_col2=get_column_letter(min_col + 9 + block_width),
            min_row=min_row + 1,
            max_col1=get_column_letter(min_col + 4 + block_width),
            max_col2=get_column_letter(min_col + 13 + block_width),
            max_row=min_row + block_height - 1)

        formula = '({minor_version_number}<>0)*({name_col}{row}<>' \
                  'INDIRECT(ADDRESS(ROW({name_col}{row}),COLUMN({name_col}{row})' \
                  ',4,,CONCATENATE("V",{major_version_number},".",{minor_version_number}-1))))'. \
            format(major_version_number=self.elda_template_parameters.MAJOR_VERSION_NUMBER_FIXED_CELL,
                   minor_version_number=self.elda_template_parameters.MINOR_VERSION_NUMBER_FIXED_CELL,
                   name_col=get_column_letter(min_col + 1 + block_width),
                   row=min_row + 1)

        self.sheet.conditional_formatting.add(cell_range,
                                              FormulaRule(formula=[formula],
                                                          stopIfTrue=True,
                                                          fill=CHANGED_DATA_FILL,
                                                          border=CHANGED_DATA_BORDER,
                                                          font=CHANGED_DATA_FONT))

        # - Empty values
        cell_range = "${min_col}${min_row}:${max_col}${max_row}".format(
            min_col=get_column_letter(min_col + 1 + block_width + 1),
            min_row=min_row + 1,
            max_col=get_column_letter(min_col + 1 + block_width + 1),
            max_row=min_row + block_height - 1)

        formula = 'AND({name_col}{row}<>"",{value_col}{row}="")'.format(
            name_col=get_column_letter(min_col + 1 + block_width),
            row=min_row + 1,
            value_col=get_column_letter(min_col + 1 + block_width + 1))

        self.sheet.conditional_formatting.add(cell_range,
                                              FormulaRule(formula=[formula],
                                                          stopIfTrue=True,
                                                          fill=MISSING_DATA_FILL))

        # - Empty parameter level
        process_level_field_index = 12 if parameter_type == 'input' else 8

        cell_range = "${min_col}${min_row}:${max_col}${max_row}".format(
            min_col=get_column_letter(min_col + process_level_field_index + block_width + 1),
            min_row=min_row + 1,
            max_col=get_column_letter(min_col + process_level_field_index + block_width + 1),
            max_row=min_row + block_height - 1)

        formula = 'AND({name_col}{row}<>"",{value_col}{row}="")'.format(
            name_col=get_column_letter(min_col + 1 + block_width),
            row=min_row + 1,
            value_col=get_column_letter(min_col + process_level_field_index + block_width + 1))

        self.sheet.conditional_formatting.add(cell_range,
                                              FormulaRule(formula=[formula],
                                                          stopIfTrue=True,
                                                          fill=MISSING_DATA_FILL))

        # - Review data
        # - - Hide review code if first version
        cell_range = "${min_col}${min_row}:${max_col}${max_row}".format(
            min_col=get_column_letter(min_col + 5 + block_width + 1),
            min_row=min_row + 1,
            max_col=get_column_letter(min_col + 5 + block_width + 1),
            max_row=min_row + block_height - 1)

        formula = '{}=0'.format(self.elda_template_parameters.MINOR_VERSION_NUMBER_FIXED_CELL)

        self.sheet.conditional_formatting.add(cell_range, FormulaRule(formula=[formula],
                                                                      stopIfTrue=True,
                                                                      font=REVIEW_HIDDEN_FONT))

        # - - Hide review code if no data
        cell_range = cell_range

        formula = '{name_col}{row}=""'.format(
            name_col=get_column_letter(min_col + 1 + block_width),
            row=min_row + 1)

        self.sheet.conditional_formatting.add(cell_range, FormulaRule(formula=[formula],
                                                                      stopIfTrue=True,
                                                                      font=REVIEW_HIDDEN_FONT))

        # - - Color according to review code
        cell_range = "${min_col}${min_row}:${max_col}${max_row}".format(
            min_col=get_column_letter(min_col + 5 + block_width + 1),
            min_row=min_row + 1,
            max_col=get_column_letter(min_col + 7 + block_width + 1),
            max_row=min_row + block_height - 1)

        formula = '${review_col}{review_row}='.format(
            review_col=get_column_letter(min_col + 5 + block_width + 1),
            review_row=min_row + 1,
        )

        self.sheet.conditional_formatting.add(cell_range,
                                              FormulaRule(formula=[formula + '0'],
                                                          stopIfTrue=True,
                                                          font=REVIEW_0_FONT))
        self.sheet.conditional_formatting.add(cell_range,
                                              FormulaRule(formula=[formula + '1'],
                                                          stopIfTrue=True,
                                                          font=REVIEW_1_FONT))
        self.sheet.conditional_formatting.add(cell_range,
                                              FormulaRule(formula=[formula + '2'],
                                                          stopIfTrue=True,
                                                          font=REVIEW_2_FONT))

        if parameter_type == 'input':
            # - Uncertainty data
            # - - SD2 or 2SD
            # - - - Grayed
            cell_range = "${min_col}${min_row}:${max_col}${max_row}".format(
                min_col=get_column_letter(min_col + 9 + block_width + 1),
                min_row=min_row + 1,
                max_col=get_column_letter(min_col + 9 + block_width + 1),
                max_row=min_row + block_height - 1)

            formula = 'AND(${name_col}{row}<>"",OR(${un_type_col}{row}="",${un_type_col}{row}="Triangle",${un_type_col}{row}="Uniform"))'.format(
                name_col=get_column_letter(min_col + block_width + 1),
                un_type_col=get_column_letter(min_col + 8 + block_width + 1),
                row=min_row + 1)

            self.sheet.conditional_formatting.add(cell_range, FormulaRule(formula=[formula],
                                                                          stopIfTrue=True,
                                                                          fill=BLOCKED_DATA_FILL))

            # - - - Needed
            cell_range = cell_range

            formula = 'AND({data_col}{row}="",OR(${un_type_col}{row}="Lognormal",${un_type_col}{row}="Normal"))'.format(
                data_col=get_column_letter(min_col + 9 + block_width + 1),
                un_type_col=get_column_letter(min_col + 8 + block_width + 1),
                row=min_row + 1)

            self.sheet.conditional_formatting.add(cell_range, FormulaRule(formula=[formula],
                                                                          stopIfTrue=True,
                                                                          fill=MISSING_DATA_FILL))

            # - - Min and Max
            # - - - Grayed
            cell_range = "${min_col}${min_row}:${max_col}${max_row}".format(
                min_col=get_column_letter(min_col + 10 + block_width + 1),
                min_row=min_row + 1,
                max_col=get_column_letter(min_col + 11 + block_width + 1),
                max_row=min_row + block_height - 1)

            formula = 'AND(${name_col}{row}<>"",OR(${un_type_col}{row}="",${un_type_col}{row}="Lognormal",${un_type_col}{row}="Normal"))'.format(
                name_col=get_column_letter(min_col + block_width + 1),
                un_type_col=get_column_letter(min_col + 8 + block_width + 1),
                row=min_row + 1)

            self.sheet.conditional_formatting.add(cell_range, FormulaRule(formula=[formula],
                                                                          stopIfTrue=True,
                                                                          fill=BLOCKED_DATA_FILL))

            # - - - Needed
            cell_range = cell_range

            formula = 'AND({data_col}{row}="",OR(${un_type_col}{row}="Triangle",${un_type_col}{row}="Uniform"))'.format(
                data_col=get_column_letter(min_col + 10 + block_width + 1),
                un_type_col=get_column_letter(min_col + 8 + block_width + 1),
                row=min_row + 1)

            self.sheet.conditional_formatting.add(cell_range, FormulaRule(formula=[formula],
                                                                          stopIfTrue=True,
                                                                          fill=MISSING_DATA_FILL))

        # Adding data validations
        # - Review
        cell_range = "${min_col}${min_row}:${max_col}${max_row}".format(
            min_col=get_column_letter(min_col + 5 + block_width + 1),
            min_row=min_row + 1,
            max_col=get_column_letter(min_col + 5 + block_width + 1),
            max_row=min_row + block_height - 1)

        validation = [x for x in self.sheet.data_validations.dataValidation
                      if x.formula1 == '"0,1,2"'][0]
        validation.add(cell_range)

        # - Uncertainty type
        if parameter_type == 'input':
            cell_range = "${min_col}${min_row}:${max_col}${max_row}".format(
                min_col=get_column_letter(min_col + 8 + block_width + 1),
                min_row=min_row + 1,
                max_col=get_column_letter(min_col + 8 + block_width + 1),
                max_row=min_row + block_height - 1)

            validation = [x for x in self.sheet.data_validations.dataValidation
                          if x.formula1 == '"Lognormal,Normal,Triangle,Uniform"'][0]
            validation.add(cell_range)

        # - Parameter level
        cell_range = "${min_col}${min_row}:${max_col}${max_row}".format(
            min_col=get_column_letter(min_col + process_level_field_index + block_width + 1),
            min_row=min_row + 1,
            max_col=get_column_letter(min_col + process_level_field_index + block_width + 1),
            max_row=min_row + block_height - 1)

        validation = [x for x in self.sheet.data_validations.dataValidation
                      if x.formula1 == '" ,Process,Project,Database"'][0]
        validation.add(cell_range)

        # - Formula
        if parameter_type == 'calculated':
            cell_range = "${min_col}${min_row}:${max_col}${max_row}".format(
                min_col=get_column_letter(min_col + 1 + block_width + 1),
                min_row=min_row + 1,
                max_col=get_column_letter(min_col + 1 + block_width + 1),
                max_row=min_row + block_height - 1)

            validation = [x for x in self.sheet.data_validations.dataValidation
                          if x.formula1 == '_xlfn.ISFORMULA(' +
                          self.elda_template_parameters.FIRST_CALCULATED_PARAMETER_VALUE_CELL + ')'][0]
            validation.add(cell_range)

        # - Numeric values
        if parameter_type == 'input':
            # - - Value
            cell_range = "${min_col}${min_row}:${max_col}${max_row}".format(
                min_col=get_column_letter(min_col + 1 + block_width + 1),
                min_row=min_row + 1,
                max_col=get_column_letter(min_col + 2 + block_width + 1),
                max_row=min_row + block_height - 1)

            validation = [x for x in self.sheet.data_validations.dataValidation
                          if
                          x.formula1 == 'NOT(_xlfn.ISFORMULA(' +
                          self.elda_template_parameters.FIRST_INPUT_PARAMETER_VALUE_CELL + '))*ISNUMBER(' +
                          self.elda_template_parameters.FIRST_INPUT_PARAMETER_VALUE_CELL + ')'][0]
            validation.add(cell_range)

            # - - Uncertainty
            cell_range = "${min_col}${min_row}:${max_col}${max_row}".format(
                min_col=get_column_letter(min_col + 9 + block_width + 1),
                min_row=min_row + 1,
                max_col=get_column_letter(min_col + 11 + block_width + 1),
                max_row=min_row + block_height - 1)

            validation = [x for x in self.sheet.data_validations.dataValidation
                          if x.formula1 == 'NOT(_xlfn.ISFORMULA(' +
                          self.elda_template_parameters.FIRST_INPUT_PARAMETER_VALUE_CELL +
                          '))*ISNUMBER(' +
                          self.elda_template_parameters.FIRST_INPUT_PARAMETER_VALUE_CELL + ')'][0]
            validation.add(cell_range)

        # Updating the last parameter cell coords
        if parameter_type == 'input':
            self.last_input_parameter_cell_coords = \
                absolute_coordinate(get_column_letter(last_param_col + block_width + 1) +
                                    str(last_param_row))
            self.sheet[self.elda_template_parameters.LAST_INPUT_PARAMETER_CELL_COORDS_CELL] = \
                self.last_input_parameter_cell_coords
        else:
            self.last_calculated_parameter_cell_coords = \
                absolute_coordinate(get_column_letter(last_param_col + block_width + 1) +
                                    str(last_param_row))

            self.sheet[self.elda_template_parameters.LAST_CALCULATED_PARAMETER_CELL_COORDS_CELL] = \
                self.last_calculated_parameter_cell_coords

    def insert_parameter(self, parameter):
        """
        Inserts the parameter data in the EldaVersionWriter sheet.

        Args:
            parameter (InputParameter or CalculatedParameter):  Parameter to insert
        """
        # Getting the range of row numbers to look at
        first_row = self.elda_template_parameters.PARAMETERS_FIRST_ROW[parameter.type]
        last_row = self.elda_template_parameters.PARAMETERS_LAST_ROW[parameter.type]

        # Looping on each row to find an empty row to write data
        row = first_row
        name_column = column_index_from_string(self.elda_template_parameters.PARAMETERS_NAME_COLUMN)

        while row <= last_row:
            # If the cell isn't empty, gets to the next row
            if self.sheet.cell(row=row, column=name_column).value is not None:

                # If the end of the block have been reached, Goes to the next one
                if row == last_row:
                    # Restarts from the first row
                    row = first_row

                    # Updating the name column number
                    name_column += self.elda_template_parameters.INPUT_PARAMETERS_BLOCK_WIDTH + 1

                # If not, just gets to the next row
                else:
                    row += 1

            # If the cell is empty, writes the data in the elda and adds a named cell to the sheet
            else:

                if parameter.type == 'Input parameter':
                    fields_order = self.elda_template_parameters.INPUT_PARAMETER_FIELDS_ORDER

                    self.inserted_input_parameters += 1
                    # If all rows are full, creates a new block
                    if self.inserted_input_parameters == self.available_input_parameter_rows:
                        self.add_parameter_block('input')

                elif parameter.type == 'Calculated parameter':
                    fields_order = self.elda_template_parameters.CALCULATED_PARAMETER_FIELDS_ORDER

                    self.inserted_calculated_parameters += 1
                    # If all rows are full, creates a new block
                    if self.inserted_calculated_parameters == self.available_calculated_parameter_rows:
                        self.add_parameter_block('calculated')

                # Looping on every field to insert the data in the right cell
                for field in fields_order:

                    # Skipping the Formula field
                    if field == 'formula':
                        continue

                    column = name_column + fields_order.index(field)

                    if field in dir(parameter):
                        self.sheet.cell(row=row, column=column).value = getattr(parameter, field)
                    elif field == 'value' and parameter.type == 'Calculated parameter':
                        self.sheet.cell(row=row, column=column).value = make_excel_formula(parameter.formula)

                # Adding named cells
                value_column = openpyxl.utils.get_column_letter(name_column + fields_order.index('value'))

                add_named_cell(name=parameter.name,
                               workbook=self.elda.workbook,
                               sheet_name=self.version,
                               coordinates=value_column + str(row))

                break  # ends the loop

    def clean_data(self):
        """
        Removes parameters and flows from the EldaVersionWriter.

        Must be done before inserting new parameters and flows. Otherwise, parameters will be inserted after existing
        ones and ancient flows might be left.
        """

        # Resetting the last input and calculated parameters cells coords
        self.sheet[self.elda_template_parameters.LAST_INPUT_PARAMETER_CELL_COORDS_CELL] = \
            self.elda_template_parameters.DEFAULT_LAST_INPUT_PARAMETER_CELL_COORDS
        self.sheet[self.elda_template_parameters.LAST_CALCULATED_PARAMETER_CELL_COORDS_CELL] = \
            self.elda_template_parameters.DEFAULT_LAST_CALCULATED_PARAMETER_CELL_COORDS

        # Cleaning cells
        for cell_range in self.elda_template_parameters.CELLS_TO_CLEAN:
            reset_cells(self.sheet, cell_range)

        # Deleting cells
        max_col = self.sheet.max_column
        for cell_range in self.elda_template_parameters.CELLS_TO_DELETE:
            cell_range = cell_range.format(get_column_letter(max_col))
            delete_cells(self.sheet, cell_range)

        # Resetting cells to their default values
        for replacement_value in self.elda_template_parameters.CELLS_TO_REPLACE:
            for cell_range in self.elda_template_parameters.CELLS_TO_REPLACE[replacement_value]:
                reset_cells(self.sheet, cell_range, replacement_value)


class EldaVersionReader:
    """ Class used to read a elda version. """

    def __init__(self, elda, version):
        """
        Args:
            elda (Elda): Elda containing the version to read
            version (str): Version to read
        """
        # Loading template parameters
        elda_template_version = elda.workbook['Status'][ELDA_TEMPLATE_VERSION_CELL].value
        self.elda_template_parameters = EldaTemplateParameters(elda_template_version=elda_template_version)

        self.sheet = elda.workbook[version]

    def to_process(self):
        """
        Method used to convert a Elda version to a instance of Process

        Returns:
            Process: Process described by the elda version
        """

        # Ensuring of the date format
        try:
            version_date = datetime.strptime(self.sheet[self.elda_template_parameters.VERSION_DATE_CELL].value,
                                             '%d/%m/%Y').strftime('%d/%m/%Y')
        except ValueError:
            version_date = datetime.date(datetime.today()).strftime('%d/%m/%Y')
        except TypeError:
            version_date = datetime.date(datetime.today()).strftime('%d/%m/%Y')

        # Creating the process
        process = Process(name=self.sheet[self.elda_template_parameters.PROCESS_NAME_CELL].value,
                          date=version_date)

        # Getting process metadata
        for attribute in self.elda_template_parameters.METADATA_CELLS.items():

            if attribute[0] != 'name':
                setattr(process, attribute[0], self.sheet[attribute[1]].value)

            process.metadata_review[attribute[0]] = \
                {'comment': self.sheet[attribute[1].replace('D', 'E')].value,
                 'review_state': self.sheet[attribute[1].replace('D', 'G')].value,
                 'comment_for_reviewer': self.sheet[attribute[1].replace('D', 'F')].value,
                 'reviewer_comment': self.sheet[attribute[1].replace('D', 'I')].value}

            # Removing empty review data
            process.metadata_review[attribute[0]] = {k: v for k, v in process.metadata_review[attribute[0]].items()
                                                     if v not in (0, None)}

            if len(process.metadata_review[attribute[0]]) == 0:
                process.metadata_review.pop(attribute[0])

        # Getting review info
        for attribute in self.elda_template_parameters.VERSION_INFO_CELLS.items():
            setattr(process, attribute[0], self.sheet[attribute[1]].value)

        # Looping on every flow row to create the corresponding instance of Flow
        row_number = self.elda_template_parameters.FIRST_FLOW_ROW_NUMBER
        while True:
            row_number_str = str(row_number)

            # Breaking the loop at the first empty flow line
            if self.sheet['B' + row_number_str].value in ("Select a type", None):
                break

            # Reading flow data
            flow_data = {}
            for flow_field in self.elda_template_parameters.FLOW_FIELDS_COLUMNS.keys():
                flow_data[flow_field] = \
                    self.sheet[self.elda_template_parameters.FLOW_FIELDS_COLUMNS[flow_field] + row_number_str].value

            # Getting the original formula
            if type(flow_data['amount']) == str:
                if is_formula_using_cells_adresses(flow_data['amount']):
                    raise ExcelFormulaError(formula=flow_data['amount'])
                flow_data['amount'] = get_original_formula(flow_data['amount'])

            # Creating a Flow instance
            flow = FlowFactory.create_flow(name=flow_data['name'],
                                           type=flow_data['type'],
                                           unit=flow_data['unit'],
                                           amount=flow_data['amount'],
                                           data_source=flow_data['data_source'],
                                           library=flow_data['library'],
                                           comment=flow_data['comment'],
                                           comment_for_reviewer=flow_data['comment_for_reviewer'],
                                           review_state=flow_data['review_state'],
                                           reviewer_comment=flow_data['reviewer_comment'],
                                           uncertainty=flow_data['uncertainty'],
                                           stdev=flow_data['stdev'],
                                           min_value=flow_data['min_value'],
                                           max_value=flow_data['max_value'])

            if flow.type in PRODUCT_FLOW_TYPES:
                flow.waste_type = flow_data['waste_type']
                flow.category = flow_data['category'] or 'ELDAM'

                # Getting the original formula
                if type(flow_data['allocation']) == str:
                    if is_formula_using_cells_adresses(flow_data['allocation']):
                        raise ExcelFormulaError(formula=flow_data['allocation'])
                    flow_data['allocation'] = get_original_formula(flow_data['allocation'])
                flow.allocation = n_str(flow_data['allocation'])

            elif flow.type in TECHNOSPHERE_FLOW_TYPES:
                flow.modification_code = flow_data['modification_code']
                flow.modification_comment = flow_data['modification_comment']
                flow.relevance_code = flow_data['relevance_code']
                flow.relevance_comment = flow_data['relevance_comment']
                flow.confidence_code = flow_data['confidence_code']
                flow.confidence_comment = flow_data['confidence_comment']

            elif flow_data['type'] in BIOSPHERE_FLOW_TYPES:
                flow.compartment = flow_data['compartment']
                flow.sub_compartment = flow_data['sub_compartment']
                flow.relevance_code = flow_data['relevance_code']
                flow.relevance_comment = flow_data['relevance_comment']
                flow.confidence_code = flow_data['confidence_code']
                flow.confidence_comment = flow_data['confidence_comment']

            process.add_flow(flow)

            row_number += 1

        # Looping on every InputParameter to create the corresponding instance of InputParameter
        row_number = self.elda_template_parameters.PARAMETERS_FIRST_ROW['Input parameter']
        name_column = column_index_from_string(self.elda_template_parameters.PARAMETERS_NAME_COLUMN)

        while True:
            # If the parameter name is empty, ends the loop
            if self.sheet.cell(row=row_number, column=name_column).value is None:
                break

            # Reading input parameter data
            input_parameter_data = {}
            for param_field in self.elda_template_parameters.INPUT_PARAMETER_FIELDS_ORDER:
                input_parameter_data[param_field] = \
                    self.sheet.cell(row=row_number,
                                    column=name_column +
                                           self.elda_template_parameters.INPUT_PARAMETER_FIELDS_ORDER.index(
                                               param_field)).value

            # Creating an instance of InputParameter
            input_parameter = InputParameter(name=input_parameter_data['name'],
                                             value=input_parameter_data['value'],
                                             comment=input_parameter_data['comment'],
                                             comment_for_reviewer=input_parameter_data['comment_for_reviewer'],
                                             review_state=input_parameter_data['review_state'],
                                             reviewer_comment=input_parameter_data['reviewer_comment'],
                                             uncertainty=input_parameter_data['uncertainty'],
                                             stdev=input_parameter_data['stdev'],
                                             min_value=input_parameter_data['min_value'],
                                             max_value=input_parameter_data['max_value'],
                                             level=input_parameter_data['level'] or 'Process')

            process.add_parameter(input_parameter)

            # If the end of the block have been reached, goes to the next block. Else, goes to the next line
            if row_number == self.elda_template_parameters.PARAMETERS_LAST_ROW['Input parameter']:
                row_number = self.elda_template_parameters.PARAMETERS_FIRST_ROW['Input parameter']
                name_column += self.elda_template_parameters.INPUT_PARAMETERS_BLOCK_WIDTH + 1
            else:
                row_number += 1

        # Same thing for calculated parameters
        row_number = self.elda_template_parameters.PARAMETERS_FIRST_ROW['Calculated parameter']
        name_column = column_index_from_string(self.elda_template_parameters.PARAMETERS_NAME_COLUMN)

        while True:
            # If the parameter name is empty, ends the loop
            if self.sheet.cell(row=row_number, column=name_column).value is None:
                break

            # Reading calculated parameter data
            calculated_parameter_data = {}
            for param_field in self.elda_template_parameters.CALCULATED_PARAMETER_FIELDS_ORDER:
                calculated_parameter_data[param_field] = \
                    self.sheet.cell(row=row_number,
                                    column=name_column +
                                           self.elda_template_parameters.CALCULATED_PARAMETER_FIELDS_ORDER.index(
                                               param_field)).value

            # Getting the original formula
            if is_formula_using_cells_adresses(calculated_parameter_data['value']):
                raise ExcelFormulaError(formula=calculated_parameter_data['value'])

            calculated_parameter_data['value'] = get_original_formula(calculated_parameter_data['value'])

            # Creating an instance of CalculatedParameter
            calculated_parameter = CalculatedParameter(name=calculated_parameter_data['name'],
                                                       formula=calculated_parameter_data['value'],
                                                       comment=calculated_parameter_data['comment'],
                                                       comment_for_reviewer=calculated_parameter_data[
                                                           'comment_for_reviewer'],
                                                       review_state=calculated_parameter_data['review_state'],
                                                       reviewer_comment=calculated_parameter_data['reviewer_comment'],
                                                       level=calculated_parameter_data['level'] or 'Process')

            process.add_parameter(calculated_parameter)

            # If the end of the block have been reached, goes to the next block. Else, goes to the next line
            if row_number == self.elda_template_parameters.PARAMETERS_LAST_ROW['Calculated parameter']:
                row_number = self.elda_template_parameters.PARAMETERS_FIRST_ROW['Calculated parameter']
                name_column += self.elda_template_parameters.INPUT_PARAMETERS_BLOCK_WIDTH + 1
            else:
                row_number += 1

        return process
