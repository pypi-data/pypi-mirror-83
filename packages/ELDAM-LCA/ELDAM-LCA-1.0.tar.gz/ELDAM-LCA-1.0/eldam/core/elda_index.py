import os
import pathlib

import openpyxl
import openpyxl.styles
from openpyxl.utils import column_index_from_string

from eldam.core.elda import Elda
from eldam.core.parameters import ELDA_INDEX_FILEPATH, ELDA_INDEX_COLUMNS, ELDA_INDEX_COLUMNS_TO_CENTER,\
    EldaTemplateParameters
from eldam.utils.exceptions import NotAnEldaError
from eldam.utils.observer import Notifier
from eldam.utils.elda import version_from_text


class EldaIndexer(Notifier):
    """ Class used to index all Eldas in a directory """

    def __init__(self, directory):
        """
        Args:
            directory (str): Path to the directory containing the Eldas to index
        """
        super().__init__()

        self.directory = directory
        self.elda_index = []

    def build_index(self):
        """ Scan the diretory and builds the index of Eldas"""

        # Building the file list
        filepaths = list(pathlib.Path(self.directory).rglob("*.[xX][lL][sS][mM]"))
        nb_files = len(filepaths)
        non_elda_files = 0

        # Getting Eldas metadata
        for i, filepath in enumerate(filepaths):
            # Notifying the suscribers of the evolution
            self.notify(step='Building index', current=i, total=nb_files - non_elda_files)

            filepath = str(filepath)
            try:
                elda = Elda(filepath)
            except NotAnEldaError:
                non_elda_files += 1
                continue

            process = elda.read_last_version().to_process()

            # Getting version review state
            if version_from_text(elda.last_version_number)[1] == 0:
                review_state = 'Under progress'

            # If there is no metadata review or less than metadata fields, then at least a data is invalid
            elif (len(process.metadata_review) == 0) or \
                    (len(process.metadata_review) < (len(EldaTemplateParameters().METADATA_CELLS))):
                review_state = 'Major corrections'
            else:
                invalid = 0
                needs_change = 0

                if process.inventory_review_state == 'Invalid':
                    invalid += 1
                elif process.inventory_review_state == 'Needs change':
                    needs_change += 1

                for data in process.metadata_review.values():
                    if 'review_state' in data:
                        if data['review_state'] == 0:
                            invalid += 1
                        elif data['review_state'] == 1:
                            needs_change += 1

                for flow in process.flows:
                    if flow.review_state == 0:
                        invalid += 1
                    elif flow.review_state == 1:
                        needs_change += 1

                for parameter in process.parameters:
                    if parameter.review_state == 0:
                        invalid += 1
                    elif parameter.review_state == 1:
                        needs_change += 1

                if invalid > 0:
                    review_state = 'Major corrections'
                elif needs_change > 0:
                    review_state = 'Minor corrections'
                else:
                    review_state = 'Reviewed and valid'

            self.elda_index.append({'file_name': os.path.basename(filepath),
                                    'location': os.path.dirname(os.path.relpath(filepath, self.directory)),
                                    'process_name': process.name,
                                    'project': process.project,
                                    'comment': process.comment,
                                    'author': process.author,
                                    'contact': process.contact,
                                    'long_term_contact': process.long_term_contact,
                                    'last_version': elda.last_version_number,
                                    'last_version_date': process.date,
                                    'status': review_state,
                                    'product_flows': process.product_flows,
                                    'technosphere_flows_nb': len(process.technosphere_flows),
                                    'biosphere_flows_nb': len(process.biosphere_flows),
                                    'input_parameters_nb': len(process.input_parameters),
                                    'calculated_parameters_nb': len(process.calculated_parameters)
                                    })

    def save_index(self, filepath=None):
        """ Saves the index in an Excel spreadsheet """
        if filepath is None:
            filepath = os.path.join(self.directory, 'EldaIndex.xlsx')

        if os.path.isfile(filepath):
            raise FileExistsError("File already exists at " + filepath)

        wb = openpyxl.load_workbook(ELDA_INDEX_FILEPATH)
        ws = wb['Elda index']

        offset = 2
        for i, elda in enumerate(self.elda_index):
            self.notify(step='Writing index', current=i, total=len(self.elda_index))

            flow_offset = 0
            for name, column in ELDA_INDEX_COLUMNS.items():
                if name == 'product_flows':
                    for product in elda['product_flows']:
                        ws[column + str(i + offset + flow_offset)].value = product.name
                        flow_offset += 1

                else:
                    ws[column + str(i + offset)].value = elda[name]

            # Adding an hyperlink to the Elda in the first cell
            ws[ELDA_INDEX_COLUMNS['file_name'] + str(i + offset)].hyperlink = \
                os.path.join(self.directory, elda['location'], elda['file_name'])
            ws[ELDA_INDEX_COLUMNS['file_name'] + str(i + offset)].style = 'Hyperlink'

            # Merging cells of the same Elda
            nb_flows = len(elda['product_flows'])
            if nb_flows > 1:
                for name, column in ELDA_INDEX_COLUMNS.items():
                    if name != 'product_flows':
                        ws.merge_cells(start_row=i + offset,
                                       end_row=i + offset + flow_offset - 1,
                                       start_column=column_index_from_string(column),
                                       end_column=column_index_from_string(column))

                        ws[column + str(i + offset)].alignment = openpyxl.styles.Alignment(vertical='center')

            # Centering text in cells
            for name, column in ELDA_INDEX_COLUMNS.items():
                if name in ELDA_INDEX_COLUMNS_TO_CENTER:
                    ws[column + str(i + offset)].alignment = openpyxl.styles.Alignment(horizontal='center',
                                                                                       vertical='center')

                if name == 'comment':
                    ws[column + str(i + offset)].alignment = openpyxl.styles.Alignment(wrap_text=True)

            offset += flow_offset - 1

        wb.save(filepath)
        self.notify(step="Elda index generated", current=1, total=1)
