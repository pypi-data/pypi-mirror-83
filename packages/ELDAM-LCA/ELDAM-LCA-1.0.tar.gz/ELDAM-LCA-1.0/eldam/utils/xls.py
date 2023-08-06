""" Functions used for xls files management """

import re
from copy import copy

from openpyxl.styles import Border
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils.cell import range_boundaries, cols_from_range
from openpyxl.formula import Tokenizer

from eldam.utils.exceptions import ExcelNameError, CoordinatesFormatError, ExcelDuplicateNameError
from eldam.utils.misc import is_number
from eldam.core.parameters import EXCEL_EXTERNAL_LINK_PATTERN


def clean_row(row):
    """
    Returns a list version of the row without potential trailing empty cells

    Args:
        row (Row):

    Returns:
        list:
    """
    try:
        result = []
        for cell in reversed(row):
            if cell.value or result != []:
                result.append(cell)

        result.reverse()

    # If the row is empty, returns an empty list
    except IndexError:
        result = []

    return result


def range_contains(range_1, range_2):
    """
    Evaluates if a range contains another.

    Args:
        range_1 (str): Range to contain
        range_2 (str): Range to be contained

    Returns:
        bool:

    Examples:
        >>> range_contains('A1:F6', 'B2:D3')
        True
        >>> range_contains('B2:D3', 'A1:F6')
        False
        >>> range_contains('A1:F3', 'B2:D6')
        False
        >>> range_contains('A1:F3', 'A1:F3')
        True
    """

    bound_1 = range_boundaries(range_1)
    bound_2 = range_boundaries(range_2)

    if bound_1[0] <= bound_2[0] and bound_1[1] <= bound_2[1] and bound_1[2] >= bound_2[2] and bound_1[3] >= bound_2[3]:
        return True
    else:
        return False


def reset_cells(worksheet, cell_range, default_value=None):
    """
    Resets cells values for every cell in a range

    Args:
        worksheet (Worksheet):
        cell_range (str):
        default_value :
    """
    for column in cols_from_range(cell_range):
        for cell_coordinates in column:
            if worksheet[cell_coordinates].value != default_value:
                worksheet[cell_coordinates].value = default_value


def delete_cells(worksheet, cell_range):
    """
    Removes cells from a worksheet (deletes value, conditional formatting, data validation and cell merging)

    Args:
        worksheet (Worksheet):
        cell_range (str):
    """

    for column in cols_from_range(cell_range):
        for cell_coordinates in column:
            # Removing value
            worksheet[cell_coordinates].value = None
            # Removing style (applying the style of cell A1)
            copy_cell_style(worksheet[cell_coordinates], worksheet['A1'])

    # Removing conditional formatting
    conditional_formattings = list(worksheet.conditional_formatting._cf_rules.keys())
    for conditional_formatting in conditional_formattings:
        ranges_to_keep = [x for x in conditional_formatting.cells.ranges
                          if not range_contains(cell_range, x.coord)]

        if len(ranges_to_keep) != 0:
            conditional_formatting.cells.ranges = conditional_formatting.sqref.ranges = ranges_to_keep
        else:
            del worksheet.conditional_formatting._cf_rules[conditional_formatting]

    # Removing data validation
    for validation in worksheet.data_validations.dataValidation:
        for validation_range in validation.cells.ranges:
            if range_contains(cell_range, validation_range.coord):
                validation.cells.ranges.remove(validation_range)

    # Remove merge cells
    merge_cells_ranges = [x.coord for x in worksheet.merged_cells.ranges]
    for merged_cells_range in merge_cells_ranges:
        if range_contains(cell_range, merged_cells_range):
            worksheet.unmerge_cells(merged_cells_range)


def style_range(worksheet, cell_range, border=Border()):
    """
    Apply styles to a range of cells as if they were a single cell.

    Args:
    worksheet (Worksheet):  Excel worksheet instance
    cell_range (str): An excel range to style (e.g. A1:F20)
    border (Border): An openpyxl Border
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    rows = worksheet[cell_range]

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        left_cell = row[0]
        right_cell = row[-1]
        left_cell.border = left_cell.border + left
        right_cell.border = right_cell.border + right


def copy_border_style(worksheet, reference_cell, cell_range):
    """
    A function used to copy the border style from a cell and apply it to a range of cells

    Args:
        worksheet (WorkSheet): The worksheet to process
        reference_cell (str): The coordinates of the cell to copy the style from
        cell_range (str): The coordinates of the range of cells to apply the style to
    """

    border_style = worksheet[reference_cell].border

    style_range(worksheet=worksheet, cell_range=cell_range, border=border_style)


def copy_cell_style(reference_cell, destination_cell):
    """
    Copies the style from one cell to another

    Args:
        reference_cell (Cell):
        destination_cell (Cell):
    """

    if reference_cell.has_style:
        destination_cell.font = copy(reference_cell.font)
        destination_cell.border = copy(reference_cell.border)
        destination_cell.fill = copy(reference_cell.fill)
        destination_cell.number_format = copy(reference_cell.number_format)
        destination_cell.protection = copy(reference_cell.protection)
        destination_cell.alignment = copy(reference_cell.alignment)


def extract_variables_from_formula(formula):
    """
    Extracts variables from an Excel formula

    Args:
        formula (str): Formula to parse

    Returns:
        list: List of strings containing the variables

    Examples:
        >>> extract_variables_from_formula("=IF(foo=5;bar;L8)")
        ['foo', '5', 'bar', 'L8']
    """
    # Adding a "=" if not present at begining of the formula
    formula = formula if formula[0] == '=' else '=' + formula

    tok = Tokenizer(formula)
    return [x.value for x in tok.items if x.type == 'OPERAND']


def is_formula_using_cells_adresses(formula):
    """
    Checks if a formula is using cells adresses and not only defined names or numbers

    Args:
        formula (str): Formula to check

    Returns:
        bool:

    Example:
        >>> is_formula_using_cells_adresses('=G45+12')
        True
        >>> is_formula_using_cells_adresses('=G45-foo')
        True
        >>> is_formula_using_cells_adresses('=foo*bar')
        False
        >>> is_formula_using_cells_adresses('=8/foo')
        False
        >>> is_formula_using_cells_adresses('=A')
        False
    """
    variables = extract_variables_from_formula(formula)
    for variable in variables:
        if re.match(r'^[a-zA-Z]{1,3}\d{1,7}$', variable) is not None:
            return True
    return False


def is_valid_excel_name(name):
    """
    Checks if a name is valid under Excel named ranges rules.

    The name should begin with a letter or underscore and should not be like excel cells coordinates or C, c, R or r.
    The name should only be composed by digits, letters and underscores.

    Example:
        >>> is_valid_excel_name('C')
        False
        >>> is_valid_excel_name('c')
        False
        >>> is_valid_excel_name('R')
        False
        >>> is_valid_excel_name('r')
        False
        >>> is_valid_excel_name('CB56')
        False
        >>> is_valid_excel_name('CB')
        True
        >>> is_valid_excel_name('FFFDDFDFF56985')
        True
        >>> is_valid_excel_name('a name with spaces')
        False
        >>> is_valid_excel_name('a_name_with_special_characters?.:!/§')
        False
        >>> is_valid_excel_name('8_name_beginning_by_a_digit')
        False
        >>> is_valid_excel_name('name_with_digit_and_underscores_54_36')
        True
        >>> is_valid_excel_name('_name_beginning_by_underscore')
        True

    Args:
        name (str):

    Returns:
        bool:
    """

    is_valid = True

    # Asserts the name doesn't contain forbiden characters
    if [c for c in name if not re.match(r'\w|_', c) is not None]:
        is_valid = False

    # Asserts the name doesn't begin with a digit.
    elif re.match(r"^\d+.*$", name) is not None:
        is_valid = False

    # Asserts the name isn't in c, C, r or R
    elif name in ('c', 'C', 'r', 'R'):
        is_valid = False

    # Asserts the name isn't formatted like excel coordinates
    elif re.match(r'^[a-zA-Z]{1,3}\d{1,7}$', name) is not None:
        numeric_part = re.search(r'\d+$', name)[0]

        # If there are only zeros in the numeric part, the name is valid
        if int(numeric_part) != 0:
            is_valid = False

    return is_valid


def try_to_make_valid_excel_name(name):
    """
    Asserts a name respects Excel's defined names rules and modifies it if not.

    Warnings:
        Some names invalidity aren't handled by this function (names with spaces or special characters)

    Args:
        name (str): Name to make valid

    Returns:
        str: Valid name

    Examples:
        >>> try_to_make_valid_excel_name('E8')
        '_E8'
        >>> is_valid_excel_name(try_to_make_valid_excel_name('a name with spaces'))
        False
    """

    if not is_valid_excel_name(name):
        name = '_' + name

    return name


def get_original_name(name):
    """
    Rollbacks changes made by try_to_make_valid_excel_name().

    Args:
        name (str): Name processed by try_to_make_valid_excel_name()

    Returns:
        str: Original name

    Examples:
        >>> get_original_name('_E8')
        'E8'
        >>> get_original_name('_valid_name')
        '_valid_name'
        >>> get_original_name('Name')
        'Name'
    """
    if name[0] == '_':
        if not is_valid_excel_name(name[1:]):
            return name[1:]

    return name


def add_named_cell(name, workbook, sheet_name, coordinates, overwrite=True):
    """
    Adds a defined name to a sheet

    Args:
        name (str): Name to attribute to the cell
        workbook (Workbook): Workbook to use
        sheet_name (str): Name of the sheet's cell
        coordinates (str): Coordinates of the cell
        overwrite (bool): Should existing parameters with same name and scope be overwitten or raise an error?
    """
    original_name = name

    # Trying to make the name respect Excel's defined names rules
    name = try_to_make_valid_excel_name(name)

    # If it is still invalid, raise an exception
    if not is_valid_excel_name(name):
        raise ExcelNameError(f"{original_name} is not a usable name for a Excel defined name.")

    if re.match(r"^[A-Z]+\d+$", coordinates):
        # Putting dollars in the coordinates
        coordinates = re.sub(r"^([A-Z]+)(\d+)$", r"$\1$\2", coordinates)
    else:
        raise CoordinatesFormatError(f"The following coordinates are not well formatted: {coordinates}")

    # localSheetId specifies that the named parameter is linked to a worksheet (allows to have several defined names
    # with the same name as long as they are in different worksheets.
    defined_name = DefinedName(name=name,
                               attr_text=f'{sheet_name}!{coordinates}:{coordinates}',
                               localSheetId=workbook.sheetnames.index(sheet_name))
    if (defined_name.name, defined_name.localSheetId) in [(x.name, x.localSheetId) for x in
                                                          workbook.defined_names.definedName]:
        if overwrite:
            workbook.defined_names.definedName = [x for x in workbook.defined_names.definedName if
                                                  (x.name, x.localSheetId) !=
                                                  (defined_name.name, defined_name.localSheetId)]
        else:
            raise ExcelDuplicateNameError("A defined name already exists with this name and scope."
                                          f"\nName: {defined_name.name}\nScope:{defined_name.localSheetId}")

    workbook.defined_names.append(defined_name)


def make_excel_formula(formula):
    """
    Makes an Excel formula from a SimaPro formula string:
        - Adds an "=" sign before the formula
        - Replaces ',' by '.'
        - Replaces invalid defined names by valid ones
        - Adds an "=" sign before Excel link

    Args:
        formula (str): Formula to process

    Returns:
        str: Excel ready formula

    Examples:
        >>> make_excel_formula('A+B')
        '=A+B'
        >>> make_excel_formula('A + 3,6')
        '=A + 3.6'
        >>> make_excel_formula('CH4 / NO2')
        '=_CH4 / _NO2'
        >>> make_excel_formula('CH_4 / NO2')
        '=CH_4 / _NO2'
        >>> make_excel_formula('2.36e-5 + CH4')
        '=2.36e-5 + _CH4'
        >>> make_excel_formula('CH4 + 2.36E-5*5.0E6')
        '=_CH4 + 2.36E-5*5.0E6'
        >>> make_excel_formula("'Path/to/file[value_to_be_linked.xlsx]Feuil1'!$C$4")
        "='Path/to/file[value_to_be_linked.xlsx]Feuil1'!$C$4"
        >>> make_excel_formula(None)

    """

    if formula is None:
        return None

    if re.match(EXCEL_EXTERNAL_LINK_PATTERN, formula) is not None:
        return '=' + formula

    # Replacing ',' by '.'
    formula = formula.replace(',', '.')

    # Extracting variables from the formula
    # Removing numbers (simple and scientific format)
    formula_to_parse = ' ' + formula  # Just a trick to recognize numbers with the following regex if they are the first in the formula
    formula_to_parse = re.sub(r'(?<=\W)\d+(\.\d+E|e\-?\d+)?(?=[\ \-\+\*\/\)\(]|$)', '', formula_to_parse)
    variables = re.split("([()+\-\/*^=<> ])", formula_to_parse)

    # Looping on every variable to check if it is valid according to Excel's defined names rules, and replacing it if
    # not.
    for variable in variables:
        if (not is_valid_excel_name(variable)) and (not is_number(variable)) and (variable not in r'()+\-\/*^=<> '):
            formula = formula.replace(variable, try_to_make_valid_excel_name(variable))

    return '=' + formula


def get_original_formula(formula):
    """
    Rollbacks modifications made by make_excel_formula

    Args:
        formula (str): Formula processed by make_excel_formula

    Returns:
        str: Original formula

    Examples:
        >>> get_original_formula('=A+B')
        'A+B'
        >>> get_original_formula('=_CH4 / _NO2')
        'CH4 / NO2'
        >>> get_original_formula('3.5')
        '3.5'
    """

    if formula[0] == '=':
        formula = formula[1:]

    # Parsing the formula
    variables = extract_variables_from_formula(formula)

    for variable in variables:
        if (not is_number(variable)) and (variable not in r'()+\-\/*^=<> '):
            formula = formula.replace(variable, get_original_name(variable))

    return formula


def compare_workbook(workbook_1, workbook_2, cells_summary=True):
    """
    Function used to spot differences between two workbooks. Returns None if no differences exists

    This function only compares values and defined names, not styles or other cells attributes.

    Args:
        workbook_1 (Workbook):
        workbook_2 (Workbook):
        cells_summary (bool): If True, returns only cell coordinates and values, else complete cell object.

    Returns:
        list or Nonetype: List of pair of cells that differs from a workbook to another
    """

    # CELL VALUES
    # Building a set with every cell coordinates of each workbook
    cells = set()

    for workbook in workbook_1, workbook_2:
        for sheet in workbook:
            for cell in sheet._cells.values():
                cells.add((sheet.title, cell.coordinate))

    # Looping on every cell to check if they exist and have the same value in both workbooks
    differences = []

    for cell in cells:
        try:
            cell_1 = workbook_1[cell[0]][cell[1]]
        except (ValueError, KeyError):
            cell_1 = None

        try:
            cell_2 = workbook_2[cell[0]][cell[1]]
        except (ValueError, KeyError):
            cell_2 = None

        if cells_summary:
            if cell_1 is not None and cell_2 is not None:
                if cell_1.value != cell_2.value:
                    differences.append((f'{cell_1.coordinate}: {cell_1.value}',
                                        f'{cell_2.coordinate}: {cell_2.value}'))
            else:
                differences.append((f'{cell_1.coordinate}: {cell_1.value}',
                                    f'{cell_2.coordinate}: {cell_2.value}'))
        else:
            if cell_1 is not None and cell_2 is not None:
                if cell_1.value != cell_2.value:
                    differences.append((cell_1, cell_2))
            else:
                differences.append((cell_1, cell_2))

    # DEFINED NAMES
    # Building a set with every defined name of each workbook
    names = set()

    for workbook in workbook_1, workbook_2:
        for name in workbook.defined_names.definedName:
            names.add(name.name)

    # Looping on every defined names to check if they exist and have the same value in both workbooks
    for name in names:
        name1 = [x for x in workbook_1.defined_names.definedName if x.name == name]
        name2 = [x for x in workbook_2.defined_names.definedName if x.name == name]

        if name1 == []:
            differences.append((None, name2[0]))
        elif name2 == []:
            differences.append((name1[0], None))
        elif (name1[0].name, name1[0].attr_text, name1[0].localSheetId) != \
                (name2[0].name, name2[0].attr_text, name2[0].localSheetId):
            differences.append((name1[0], name2[0]))

    if len(differences) == 0:
        differences = None

    return differences
