"""
Cells styles used in the Elda
"""

from openpyxl.styles import PatternFill, Color, Font, Border, Side

CHANGED_DATA_FILL = PatternFill(bgColor=Color(rgb='FFFFFF99'))
CHANGED_DATA_FONT = Font(bold=True)
CHANGED_DATA_BORDER = Border(left=Side(border_style='thin', color='FFF79646'),
                             right=Side(border_style='thin', color='FFF79646'),
                             top=Side(border_style='thin', color='FFF79646'),
                             bottom=Side(border_style='thin', color='FFF79646'))

MISSING_DATA_FILL = PatternFill(bgColor=Color(tint=0.3999450666829432, theme=5))

BLOCKED_DATA_FILL = PatternFill(fill_type='darkTrellis')

REVIEW_0_FONT = Font(color=Color(rgb='FFC00000'))
REVIEW_1_FONT = Font(color=Color(theme=9, tint=-0.249946592608417))
REVIEW_2_FONT = Font(color=Color(rgb='FF0070C0'))
REVIEW_HIDDEN_FONT = Font(color=Color(rgb='FFFFFFFF'))
