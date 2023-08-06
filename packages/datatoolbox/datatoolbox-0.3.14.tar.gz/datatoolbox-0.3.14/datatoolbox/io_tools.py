#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Mar 19 09:58:27 2019

@author: Andreas Geiges
"""
import re
import os
import string
import pandas as pd
import numpy as np
from . import config

from . import core
from . import util
from . import mapping as mapp

from . import greenhouse_gas_database as gh
GHG_data = gh.GreenhouseGasTable()
from .data_structures import Datatable,TableSet
from openpyxl import load_workbook

REQUIRED_SETUP_FIELDS = ['filePath',
                         'fileName',
                         'sheetName',
                         'timeIdxList',
                         'spaceIdxList']

SP_ARG = 3
TI_ARG = 2
DT_ARG = 4



#%% Functions
def alphaCol2Num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) +1 
    return num-1

def colNameToNum(name):
    pow = 1
    colNum = 0
    for letter in name[::-1]:
            colNum += (int(letter, 36) -9) * pow
            pow *= 26
    return colNum -1

def excelIdx2PandasIdx(index):
    [alpha, num, _] = re.split(r'(\d+)',index)    
    return (int(num)-1, alphaCol2Num(alpha))



def getDefaultSetup():
    print("""
setup = dict()
setup['filePath']  = 'path/to/template/'
setup['fileName']  = 'template.xlsx'
setup['sheetName'] = 'templateSheet0'
setup['timeIdxList']  = ('B1', 'C5')
setup['spaceIdxList'] = ('A2', 'A20')
    """) 


def read_MAGICC6_MATLAB_bulkout(pathName):
    
    temp_offset = 0.61
    df = pd.read_table(pathName, skiprows=23, header=0, delim_whitespace=True, index_col=0)
    df = df + temp_offset
    df.index = df.index.astype(int)
    return df
    #df.values[df.values<0] = np.nan
#    
#    meta = dict()
#    meta['entity'] = 'global_temp'
#    meta['unit']   = '°C'
#    return Datatable(df, meta=meta)

def read_MAGICC6_BINOUT(filePath):
    import pymagicc as pym
    reader = pym.io._BinaryOutReader(filePath)

    metaData, df = reader.read()
    
    data = df.pivot(index='region',columns='time', values='value')
    
    metaDict = dict()
    metaDict['entity'] = df.variable[0]
    metaDict['source'] = 'MAGICC6_calculation'
    metaDict['unit']   = None

    return Datatable(data, meta=metaDict)

def read_PRIMAP_csv(fileName):
    
    metaMapping = {
            'entity': 'SHEET_ENTITY',
            'unit':'SHEET_UNIT',
            'category':'SHEET_NAME_CATEGORY',
            'scenario': 'SHEET_SCENARIO',
            'model' : 'SHEET_SOURCE'
            }
    
    allDf = pd.read_csv(fileName, usecols=[0,1], index_col=0)
    #print(os.path.basename(fileName))

    firstDataRow = allDf.loc['SHEET_FIRSTDATAROW','Unnamed: 1']
    
    #bugfix of old wrong formated PRIMAP files
    try:
        int(firstDataRow)
    except:
        firstDataRow = pd.np.where(allDf.index == "Countries\Years")[0][0]+3
   
    firstMetaRow = pd.np.where(allDf.index == "&SHEET_SPECIFICATIONS")[0][0]+1
    
    metaPrimap = dict()
    for row in range(firstMetaRow,firstDataRow):
        key = allDf.index[row]
        value = allDf.iloc[row,0]
        if key =='/':
            break
        metaPrimap[key] = value
    
    data = pd.read_csv(fileName, header=firstDataRow-2, index_col=0)
    
    meta = dict()
    for metaKey in metaMapping:
        
        if isinstance(metaMapping[metaKey], list):
            value = '_'.join(metaPrimap[x] for x in metaMapping[metaKey])
        else:
            value = metaPrimap[metaMapping[metaKey]]
        meta[metaKey] = value
    
    
    
    table = Datatable(data, meta=meta)
    table = table.loc[:, util.yearsColumnsOnly(table)]
    table.columns = table.columns.astype(int)
    return table#, data

def read_PRIMAP_Excel(fileName, sheet_names= None, xlsFile=None):
    if sheet_names is None:
        sheet_names = getAllSheetNames(fileName)
        
    out = TableSet()
    for sheet_name in sheet_names:
        table = _read_PRIMAP_Excel_single(fileName, sheet_name, xlsFile=xlsFile)
        out[sheet_name] = table
        
    return out
        
def _read_PRIMAP_Excel_single(fileName, sheet_name= 0, xlsFile=None):
    if xlsFile is not None:
        xlsFile = xlsFile
    else:   
        xlsFile = pd.ExcelFile(fileName)
    allDf = pd.read_excel(xlsFile, sheet_name = sheet_name, usecols=[0,1], index_col=0)
    #print(os.path.basename(fileName))

    firstDataRow = allDf.loc['SHEET_FIRSTDATAROW','Unnamed: 1']
    
    #bugfix of old wrong formated PRIMAP files
    try:
        int(firstDataRow)
    except:
        firstDataRow = pd.np.where(allDf.index == "Countries\Years")[0][0]+3
        
    #print(firstDataRow)
    setup = dict()
    setup['filePath']  = os.path.dirname(fileName) +'/'
    setup['fileName']  = os.path.basename(fileName)
    setup['sheetName'] = sheet_name
    setup['timeIdxList']  = ('B' + str(firstDataRow-1), 'XX'+ str(firstDataRow-1))
    setup['spaceIdxList'] = ('A'+ str(firstDataRow), 'A1000')
    #print(setup)
    ex = ExcelReader(setup, xlsFile=xlsFile)
    data = ex.gatherData().astype(float)
    #return data
    meta = {'source': '',
            'entity': allDf.loc['SHEET_ENTITY','Unnamed: 1'],
            'unit': allDf.loc['SHEET_UNIT','Unnamed: 1'],
            'category': allDf.loc['SHEET_NAME_CATEGORY','Unnamed: 1'],
            'scenario': allDf.loc['SHEET_SCENARIO','Unnamed: 1'] + '|' + allDf.loc['SHEET_SOURCE','Unnamed: 1']}
    REG_ton = re.compile('^[GM]t')
    xx = REG_ton.search(meta['unit'])
    
    if xx:
        meta['unit'] = meta['unit'].replace(xx.group(0), xx.group(0) + ' ')
    
    table = Datatable(data, meta=meta)
    try:
        table = table.loc[:,util.yearsColumnsOnly(table)]
        table.columns = table.columns.astype(int)
#        table = table.loc[:,util.yearsColumnsOnly(table)]
    except:
        print('warning: Columns could not be converted to int')
    return table

def read_MAGICC6_ScenFile(fileName, **kwargs):
    VALID_MASS_UNITS= {
            'Pt': 1e18,
            'Gt': 1e15,
            'Mt': 1e12,
            'kt': 1e9,
            't' : 1e6,
            'Pg': 1e15,
            'Tg': 1e12,
            'Gg': 1e9,
            'Mg': 1e6,
            'kg': 1e3,
            'g' : 1}
    fid = open(fileName,'r')
    nDataRows    = int(fid.readline().replace('/n',''))
    
    while True:
    #for i, line in enumerate(fid.readlines()):
        line = fid.readline()
        if line[:11] == '{0: >11}'.format('YEARS'):
            break
    #get first header line
    entities  = line.split()[1:]
    
    
    #reading units
    unitLine = fid.readline().split()[1:]
    #print(unitLine)
    
    # find correct component
    components = [GHG_data.findEntryIdx(entity) for entity in entities]

    units = [unit for unit in unitLine if unit[:2] in VALID_MASS_UNITS]        
    
    
    replacementDict = {'MtN2O-N' : 'Mt N'}

    units = [replacementDict.get(unit, unit) for unit in units]

    columns = [(x,y,z) for x,y,z in zip(entities, components, units)]
    entityFrame =  pd.DataFrame(columns=entities)
    entityFrame.columns = pd.MultiIndex.from_tuples(columns)
    
    entityFrame.columns.names=['NAME', 'COMP','UNIT']

    for i, line in enumerate(fid.readlines()):
        
        if i == nDataRows:
            break
        data = line.split()
        entityFrame.loc[int(data[0])] = np.asarray(data[1:])

    # TODO: CHange to a results list of datatables
    return entityFrame

def insertDataIntoExcelFile(fileName, 
                            overwrite = False, 
                            setupSheet = 'INPUT',
                            interpolate = False):
    ins = ExcelWriter(fileName=fileName, overwrite = overwrite, setupSheet = setupSheet, interpolate=interpolate)
    ins.insert_data()
    ins.close()
    return ins
    
class ExcelDatablock():
    def __init__(self):
        self.type        = 'excel'
        self.filePath    = None
        self.fileName    = None
        self.sheetName   = None
        self.timeIdxList  = (None, None)
        self.spaceIdxList = (None, None)
        self.parameters  = list(self.__dict__.keys())
    
    def toDict(self):
        out = dict()
        for para in self.parameters:
            out[para] = self.__dict__[para]
        return out
#%% Classes    
    
def getAllSheetNames(filePath):
    xlFile = pd.ExcelFile(filePath)
    sheetNameList = xlFile.sheet_names
    xlFile.close()
    return sheetNameList


class ExcelReader():
    
    def __init__(self, setupDict, xlsFile=None):
        """
        Required setup parameters:
            filePath    = './data/'
            fileName    = "test.xls"
            sheetName   = "Sheet0"
            timeIdxList  = ("B1", "C1")
            spaceIdxList = ('A2', 'A10')
        
        """
        self.setup(**setupDict)
        if xlsFile is not None:
            self.xlsFile = xlsFile
        
    
#    def setSetup(self, **kwargs):
#        for key in kwargs.keys():
#            self.setup[key] = kwargs[key]

    def setup(self, **kwargs):
        for key in kwargs.keys():
            
            
            if 'Idx' in key:
                # conversion of index
                setattr(self, key, [excelIdx2PandasIdx(x) for x in kwargs[key]])
            else:
                setattr(self, key, kwargs[key])
        if hasattr(self,'df'):
            del self.df
#           
#        self.validate()
    
    def getAllSheetNames(self):
        xlFile = pd.ExcelFile(os.path.join(self.filePath, self.fileName))
        sheetNameList = xlFile.sheet_names
        xlFile.close()
        return sheetNameList

    def validate(self,):
        for attr in REQUIRED_SETUP_FIELDS:
            if not hasattr(self, attr):
                print(attr + " is missing")

    def gatherValue(self, excelIdx):
        if not hasattr(self, 'df'):
            if hasattr(self, 'xlsFile'):
                self.df = pd.read_excel(self.xlsFile, sheet_name=self.sheetName, header=None)
            else:
                self.df = pd.read_excel(os.path.join(self.filePath, self.fileName), sheet_name=self.sheetName, header=None)
        elif core.config.DEBUG:
            print('use loaded df')
        return self.df.iloc[excelIdx2PandasIdx(excelIdx)]             

    def getAllData(self):
        return self.df
        
    def openFile(self):
        import os
        os.system('libreoffice ' + os.path.join(self.filePath, self.fileName)) 
                        
    def gatherData(self, load=True):

        if load:
            self.df = pd.read_excel(os.path.join(self.filePath, self.fileName), sheet_name=self.sheetName, header=None)
            print('df loaded')
        elif core.config.DEBUG:
            print('use loaded df')

#        print(self.spaceIdxList)
#        print(self.timeIdxList)
        
        if len(self.timeIdxList) >1:
            timeIdx = self.df.iloc[self.timeIdxList[0][0],self.timeIdxList[0][1]:self.timeIdxList[1][1]+1]
            columns = self.df.columns[self.timeIdxList[0][1]:self.timeIdxList[1][1]+1]
#            print(timeIdx)
        else:
            timeIdx = self.df.iloc[self.timeIdxList[0][0],[self.timeIdxList[0][1]]]
            columns = self.df.columns[[self.timeIdxList[0][1]]]

        if len(self.spaceIdxList) >1:
            regions   = self.df.iloc[self.spaceIdxList[0][0]:self.spaceIdxList[1][0]+1,self.spaceIdxList[0][1]]
            index = self.df.index[self.spaceIdxList[0][0]:self.spaceIdxList[1][0]+1]
        else:
            regions   = self.df.iloc[[self.spaceIdxList[0][0]],self.spaceIdxList[0][1]]
            index = self.df.index[[self.spaceIdxList[0][0]]]
            
        if len(index) == 1:
#            print(index)
#            print(columns)
#            print(regions)
#            print(timeIdx)

            if len(columns) == 1:
                # extraction of a single value    
                data    = self.df.loc[index,columns]
                return pd.DataFrame(data.values, columns=regions, index=timeIdx).T
            
            else:
                # extraction of time serie
                data    = self.df.iloc[self.spaceIdxList[0][0],self.timeIdxList[0][1]:self.timeIdxList[1][1]+1]
                #print(type(data))
                #print(data.index)
                return pd.DataFrame(data.values, columns = regions, index = timeIdx).T
#                return pd.Series(data.values, index=columns) 
            
        else:
        
            if len(columns) == 1:
                # extraction of spatial series
                data    = self.df.loc[index,columns]
                
                return pd.DataFrame(data.values, columns=regions, index=timeIdx).T
#                return pd.Series(data.values, index=index) 
            else:
                # extraction of both spatial and time data
#                print(self.df)
#                print(index)
#                print(timeIdx)
#                print(columns)
                data    = self.df.loc[index,columns]
#                print(data)
                
                return pd.DataFrame(data.values, columns=timeIdx, index=regions)



REG_EXCEL_RANGE = re.compile('^[A-Z]{1,3}[0-9]{1,3}:[A-Z]{1,3}[0-9]{1,3}$')
REG_EXCEL_ROW   = re.compile('^[0-9]{1,2}$')
REG_EXCEL_COL   = re.compile('^[A-Z]{1,2}$')

REG_FIND_ROWS   = re.compile('^[A-Z]{1,3}([0-9]{1,3}):[A-Z]{1,3}([0-9]{1,3})$')
REG_FIND_COLS   = re.compile('^([A-Z]{1,3})[0-9]{1,3}:([A-Z]{1,3})[0-9]{1,3}$')

def isColRange(string):
#    if REG_EXCEL_COL.match(string):
#        return True
    match = REG_FIND_COLS.search(string)
    if match:
        if match.group(1) == match.group(2):
            return True
    
    return False
        
def isRowRange(string):
#    if REG_EXCEL_ROW.match(string):
#        return True
    match = REG_FIND_ROWS.search(string)
    if match:
        if match.group(1) == match.group(2):
            return True
    
    return False

def isRow(string):
    if REG_EXCEL_ROW.match(string):
        return True
    else:
        return False
    
def isCol(string):
    if REG_EXCEL_COL.match(string):
        return True
    else:
        return False
#%%
def iterTime(timeIdxString, wksheet, xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx):
    if isRowRange(timeIdxString):
        # colum setup
        for timeCell in wksheet[timeIdxString][0]:
            xlsCol = timeCell.col_idx
            timeIdx = timeCell.value
            #print(xlsCol, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    
    elif isRow(timeIdxString):
        for timeCell in wksheet[timeIdxString]:
            xlsCol = timeCell.col_idx
            timeIdx = timeCell.value
            #print(xlsRow, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    elif isCol(timeIdxString):
        for timeCell in wksheet[timeIdxString]:
            xlsRow = timeCell.row
            timeIdx = timeCell.value
            #print(xlsRow, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
            
    elif isColRange(timeIdxString):
        for timeCell in wksheet[timeIdxString]:
#            print(timeCell[0].row)
            xlsRow = timeCell[0].row
            timeIdx = timeCell[0].value
            #print(xlsRow, timeIdx)
#            print([xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx])
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    else:
        # assume fixed
        timeIdx = timeIdxString
        yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

def iterSpace(spaceIdxString, wksheet, xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx):
    if isRowRange(spaceIdxString):
        # colum setup
        for spaceCell in wksheet[spaceIdxString][0]:
            xlsCol = spaceCell.col_idx
            spaceIdx = spaceCell.value
            #print(xlsCol, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

    elif isRow(spaceIdxString):
        for spaceCell in wksheet[spaceIdxString]:
            xlsCol = spaceCell.col_idx
            spaceIdx = spaceCell.value
            #print(xlsRow, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    
    elif isCol(spaceIdxString):
        for spaceCell in wksheet[spaceIdxString]:
            xlsRow = spaceCell.row
            spaceIdx = spaceCell.value
            #print(xlsRow, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
            
    elif isColRange(spaceIdxString):
        for spaceCell in wksheet[spaceIdxString]:
            xlsRow = spaceCell[0].row
            spaceIdx = spaceCell[0].value
            #print(xlsRow, spaceIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    else:
        # assume fixed string
        print('fixed')
        spaceIdx = spaceIdxString
        yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
        
def iterData(dataIdxString, wksheet, xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx):

    #print([xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx])
    if isRowRange(dataIdxString):
        # colum setup
        for dataCell in wksheet[dataIdxString][0]:
            xlsCol = dataCell.col_idx
            dataIdx = dataCell.value
            #print(xlsCol, timeIdx, dataIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

    elif isRow(dataIdxString):
        for dataCell in wksheet[dataIdxString]:
            xlsCol = dataCell.col_idx
            dataIdx = dataCell.value
            #print(xlsRow, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    
    
    elif isCol(dataIdxString):
#        print('isCol')
#        print([xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx])
    
        for i, dataCell in enumerate(wksheet[dataIdxString]):
#            print(dataCell)
#            print(dataCell.col_idx)
#            xlsRow = dataCell.row
            xlsCol = dataCell.col_idx
            if i == 0:
                dataIdx = dataCell.value
            #print(xlsRow, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
            
    elif isColRange(dataIdxString):
#        print('isColRange')
        for dataCell in wksheet[dataIdxString]:
            #print(dataCell)
            xlsRow = dataCell[0].row
            if dataCell[0].value is not None:
                dataIdx = dataCell[0].value.replace('\ufeff','')
            else:
                dataIdx = None
            #print(xlsRow, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    else:
        # assume fixed
        dataIdx = dataIdxString
        yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]


def _str2float(x):
#    print(x)
    if isinstance(x,float) or isinstance(x,int) or x is None:
        return x
    if '%' in x:
        return float(x.replace('%',''))*100
    else:
        if x.startswith('#') or x == '':
            return np.nan
#        try:
        print(x)
        return float(x)
#        except:
#            print(x)
#            sdf
    
def pandasStr2floatPercent(X):
    return([_str2float(x) for x in X])        
#%%
if False:
    #%%
    ws_readValues = load_workbook(self.setup['fileName'], data_only=True)['Sheet6']
    timeIdxString  = 'B2:K2'
    spaceIdxString = 'A4:A8'
    dataIdxString  = 'population|all|PROJECTION_MED|UN_WPP2017'
    
    ws_readValues = load_workbook(self.setup['fileName'], data_only=True)['Sheet7']
    timeIdxString  = 'B1:O1'
    spaceIdxString = 'DEU'
    dataIdxString  = 'A2:A6'
    
    ws_readValues = load_workbook(self.setup['fileName'], data_only=True)['Sheet8']
    timeIdxString  = '2020'
    spaceIdxString = 'B1:O1'
    dataIdxString  = 'A2:A6'

#Sheet2	B2:E2	A3:A4	
    ws_readValues = load_workbook(self.setup['fileName'], data_only=True)['Sheet2']
    timeIdxString  = 'B2:E2'
    spaceIdxString = 'A3:A4'
    dataIdxString  = 'area_agriculture|agriculture|historic|WDI2018'
##Sheet2	B2:E2	A3:A4	
#    ws_readValues = load_workbook(self.setup['fileName'], data_only=True)['Sheet4']
#    timeIdxString  = 'A'
#    spaceIdxString = 'B2:D2'
#    dataIdxString  = 'population|all|PROJECTION_LOW|UN_WPP2017'

    ws_readValues = load_workbook(self.setup['fileName'], data_only=True)['Sheet4a']
    timeIdxString  = 'A'
    spaceIdxString = '2'
    dataIdxString  = 'population|all|PROJECTION_LOW|UN_WPP2017'

    
    args =  None, None, None, None, None
    tableIDs = list()
    for args in iterData(dataIdxString, ws_readValues, *args):
        tableIDs.append(args[4])
    tables = dt.getTables(tableIDs)
    wksSheet = ws_readValues
    args =  None, None, None, None, None
    for argsTime in iterTime(timeIdxString, wksSheet, *args):
        if argsTime[TI_ARG] is None:
            continue
        for argsSpace in iterSpace(spaceIdxString, wksSheet, *argsTime):
            if argsSpace[SP_ARG] is None:
                continue
            for argData in iterData(dataIdxString, wksSheet, *argsSpace):
                if argData[DT_ARG] is None:
                    continue
                print(argData)
                try:
                    value = tables[argData[DT_ARG]].loc[argData[SP_ARG],int(argData[TI_ARG])]
                    print('success')
                except:
                    pass
#%%
class ExcelWriter():
    """
    More complex excel interface
    Author: Andreas Geiges
    """
    def __init__(self, setup=None, 
                 fileName=None, 
                 overwrite=False, 
                 setupSheet = 'INPUT',
                 interpolate=False):
        self.overwrite = overwrite # overwriting values
        self.setupSheet= setupSheet
        if setup:
            self.setup = setup
            fileSetup = True
        else:
            self.setup = dict()
            self.setup['fileName'] = fileName
            fileSetup = False
    
        self.interpolate = interpolate
    
    def close(self):
        self.wb.close()
        
        
    def getSetups(self):
        self.wb_read = load_workbook(self.setup['fileName'], data_only=True)
        self.wb = load_workbook(self.setup['fileName'])
        setup = dict()
        setup['fileName'] = self.setup['fileName']
        
        # new using pandas
        mapping = pd.read_excel(self.setup['fileName'], sheet_name=self.setupSheet)
        mapping.columns = [x.lower() for x in mapping.columns]
        
        
        setupDict = {'sheetName'     : 'sheetname',
                     'timeIdxList'   : 'time',
                     'spaceIdxList'  : 'region',
                     'dataID'        : 'variable',
                     'unit'          : 'unit',
#                     'unitTo'        : 'unitto'
                     }
        
        for i,setupMapp in mapping.iterrows():
            print(setupMapp)
            for key in setupDict.keys():
                setup[key] = str(setupMapp[setupDict[key]])
                if setup[key] == 'nan':
                    setup[key] = None
            yield(dict(setup))
        # old 
#        setupSheet = self.wb[self.setupSheet]
#        for row in setupSheet.rows:
#            if row[0].value == 'sheet':
#                continue
#            
#            setup['sheetName']   = row[0].value
#            setup['timeIdxList']  = row[1].value
#            setup['spaceIdxList'] = row[2].value
#            setup['dataID']      = row[3].value
#            setup['unit']        = row[4].value
#            
#            for key in setup.keys():
#                if isinstance(setup[key],str):
#                    setup[key] = setup[key].replace('\ufeff','')
#            yield(setup)
        
    def insert_data(self):
        replaceDict = {'EU': 'EU28'}
        from shutil import copyfile
        from copy import copy
        
        # Copy old file with "_filled" extension and load it
        if '.xlsx' in self.setup['fileName']:
            saveFileName = self.setup['fileName'].replace('.xlsx','_filled.xlsx')
        elif '.xls' in self.setup['fileName']:
            saveFileName = self.setup['fileName'].replace('.xls','_filled.xlsx')
        else:
            print('file extention not recognized')
        copyfile(self.setup['fileName'], saveFileName)
        self.setup['fileName'] = saveFileName
#        wb = load_workbook(self.setup['fileName'])
        
        
        # pre-load all valid spatial IDS
        validSpatialIDs = mapp.getValidSpatialIDs()
        
        # create list of setups
        self.setupList = list()
        
        #loop over all defined inputs in the INPUT sheet mapping (rows)
        for setup in self.getSetups():
            
            #ensure input is at all times a string
            if isinstance(setup['timeIdxList'], int):
                setup['timeIdxList'] = str(setup['timeIdxList'])
            if isinstance(setup['spaceIdxList'], int):
                setup['spaceIdxList'] = str(setup['spaceIdxList'])
            
            if config.DEBUG:
                print(setup)
            
            #add a copy of setup to list
            self.setupList.append(copy(setup))
            
            args =  None, None, None, None, None
            wksSheet_read  = self.wb_read[setup['sheetName']]
            wksSheet_write = self.wb[setup['sheetName']]
            # loop overall setups and collect dataIDs for pre-loading
            tableIDs = list()
            for args in iterData(setup['dataID'], wksSheet_read, *args):
                
                if config.DEBUG:
                    print(args)
                if (~pd.isna(args[4])) and (args[4] is not None):
                    tableIDs.append(args[4])
            
            if config.DEBUG:
                print(tableIDs)
                
            #load all tables from database
            tables = core.DB.getTables(tableIDs)
            
            if setup['unit'] is not None:
                 for tableKey in tables.keys():
                     table = tables[tableKey]
                     tables[tableKey]= table.convert(setup['unit'])
                     
            if self.interpolate:
                 for tableKey in tables.keys():
                     for col in list(range(tables[tableKey].columns.min(),tables[tableKey].columns.max()+1)):
                         if col not in tables[tableKey].columns:
                             tables[tableKey].loc[:,col] = np.nan
                     tables[tableKey] = tables[tableKey].interpolate()    
                     
#            wksSheet = load_workbook(setup['fileName'], data_only=True)[setup['sheetName']]
            
            args =  [None, None, None, None, None]
            iCount = 0
            
            #iterate over time index list
            for argsTime in iterTime(setup['timeIdxList'], wksSheet_read, *args):
                if argsTime[TI_ARG] is None:
                    print('No time defintion found')
                    continue
                
                #iterate over space index list
                for argsSpace in iterSpace(setup['spaceIdxList'], wksSheet_read, *argsTime):
                    if argsSpace[SP_ARG] is None:
                        print('not spacial defintion found')
                        continue
                    
                    
                    if argsSpace[SP_ARG] not in validSpatialIDs:
#                        print('not time defintion found')
#                        print(argsSpace[SP_ARG])
                        if argsSpace[SP_ARG] in replaceDict.keys():
                            newID = replaceDict[argsSpace[SP_ARG]]
                        else:
                            newID = util.getCountryISO(argsSpace[SP_ARG])
                        

                        if newID in validSpatialIDs:
                            argsSpace[SP_ARG] = newID
                            print(argsSpace[SP_ARG])
                        else:
                            print(argsSpace[SP_ARG] + ' not found')
                            continue
                        
                    #print(argsSpace[SP_ARG])
                    #iterate over all data indices
                    for argData in iterData(setup['dataID'], wksSheet_read, *argsSpace):
                        if argData[DT_ARG] is None:
                            print('not data defintion found')
                            continue

                        try:
                            value = tables[argData[DT_ARG]].loc[argData[SP_ARG],int(argData[TI_ARG])]
                            self._writeValueNew(wksSheet_write, xlsRow=argData[0], xlsCol=argData[1], value=value)
                            
#                            print('success')
                            iCount +=1
                            
                        except Exception as e:
                            if config.DEBUG:
                                print('failed with agruments:' + str(argData))
                                print(e)
                                #import pdb
                                #pdb.set_trace()
                            pass

            self.wb.save(self.setup['fileName'])
            print(setup)
            print('{} items inserted'.format(iCount))
            
            self.wb.close()

    def _writeValueNew(self, wrkSheet, xlsRow, xlsCol, value):
        if self.overwrite or pd.isna(wrkSheet.cell(row=xlsRow, column=xlsCol).value):
            print(value)
            wrkSheet.cell(row=xlsRow, column=xlsCol, value = value)
        
#    def _writeMultipleIndicators(self, setup):
#        ws_readValues = load_workbook(self.setup['fileName'], data_only=True)[setup['sheetName']]
#        try:
#            tableIDs = [cell.value for cell in ws_readValues[setup['dataID']]]
#        except:
#            tableIDs = [cell[0].value for cell in ws_readValues[setup['dataID']]]
#            
#            tables = dt.getTables(tableIDs)
#            
#            REG_FIND_ROWS.search(setup['timeIdxList'])
            

def PandasExcelWriter(fileName):
    return pd.ExcelWriter(fileName,
                        engine='xlsxwriter',
                        datetime_format='mmm d yyyy hh:mm:ss',
                        date_format='mmmm dd yyyy')


#%%
class ExcelReader_New():
    
    def __init__(self, setup=None, fileName=None, overwrite=False, setupSheet = 'OUTPUT'):
        self.overwrite = overwrite
        self.setupSheet= setupSheet
        if setup:
            self.setup = setup
            fileSetup = True
        else:
            self.setup = dict()
            self.setup['fileName'] = fileName
            fileSetup = False
    
    def close(self):
        self.wb.close()
        
    def openSourceFile(self):
        if config.OS == 'Linux':
            os.system('libreoffice ' + self.setup['fileName'])
        elif dt.conf.OS == 'Darwin':
            os.system('open -a "Microsoft Excel" ' + self.setup['fileName'])        
   
    def getSetups(self):
        self.wb_read = load_workbook(self.setup['fileName'], data_only=True)
        self.wb = load_workbook(self.setup['fileName'], data_only=True)
        setup = dict()
        setup['fileName'] = self.setup['fileName']
        
        
        mapping = pd.read_excel(self.setup['fileName'], sheet_name =self.setupSheet)
        
        
        
        setupDict = {'sheetName'     : 'sheetName',
                     'timeIdxList'   : 'timeCells',
                     'spaceIdxList'  : 'spaceCells',
                     'dataID'        : 'dataCells',
                     'unit'          : 'unit',
                     'unitTo'        : 'unitTo'}

        setupDict = {'sheetName'     : 'Sheetname',
                     'timeIdxList'   : 'Time',
                     'spaceIdxList'  : 'Region',
                     'dataID'        : 'Variable',
                     'unit'          : 'Unit',
                     'unitTo'        : 'UnitTo',
                     'scenario'      : 'Scenario',
                     'source'        : 'Source'}
        
        for i,setupMapp in mapping.iterrows():
            for key in setupDict.keys():
                setup[key] = str(setupMapp[setupDict[key]])
                if setup[key] == 'nan':
                    setup[key] = None
                    
            yield(dict(setup))
            
#        setupSheet = self.wb[self.setupSheet]
#        for row in setupSheet.rows:
#            if row[0].value == 'sheet':
#                continue
#            
#            setup['sheetName']   = row[0].value
#            setup['timeIdxList']  = row[1].value
#            setup['spaceIdxList'] = row[2].value
#            setup['dataID']      = row[3].value
#            setup['unit']        = row[4].value
#            
#            for key in setup.keys():
#                if isinstance(setup[key],str):
#                    setup[key] = setup[key].replace('\ufeff','')
#            yield(setup)
        
    def read_data(self):
        replaceDict = {'EU': 'EU28'}
        from shutil import copyfile
        from copy import copy

        wb = load_workbook(self.setup['fileName'], data_only=True)
        
        validSpatialIDs = mapp.getValidSpatialIDs()
        self.setupList = list()
        
        tablesToReturn = TableSet()
        
        for setup in self.getSetups():
            if isinstance(setup['timeIdxList'], int):
                setup['timeIdxList'] = str(setup['timeIdxList'])
            if isinstance(setup['spaceIdxList'], int):
                setup['spaceIdxList'] = str(setup['spaceIdxList'])
            print(setup)
            self.setupList.append(copy(setup))
            args =  None, None, None, None, None
            wksSheet = wb[setup['sheetName']]

            
            
            args =  [None, None, None, None, None]
            for argsTime in iterTime(setup['timeIdxList'], wksSheet, *args):
                
                if argsTime[TI_ARG] is None:
                    print('No time defintion found')
                    print(setup['timeIdxList'])

                    continue
                print(argsTime[TI_ARG])
                print(setup['spaceIdxList'])
                for argsSpace in iterSpace(setup['spaceIdxList'], wksSheet, *argsTime):
                    if argsSpace[SP_ARG] is None:
                        print('Not space defintion found')
                        continue
                    
                    if argsSpace[SP_ARG] not in validSpatialIDs:
                        print('No valid ISO code...')
                        
                        
                        if argsSpace[SP_ARG] in replaceDict.keys():
                            newID = replaceDict[argsSpace[SP_ARG]]
                        else:
                            newID = util.getCountryISO(argsSpace[SP_ARG])
                        

                        if newID in validSpatialIDs:
                            argsSpace[SP_ARG] = newID
                            print('Set iso code to: ' + argsSpace[SP_ARG])
                        else:
                            print('No ISO code found')
                            print(argsSpace[SP_ARG] + ' not found')
                            continue
                        
                    print(argsSpace)
                    for argData in iterData(setup['dataID'], wksSheet, *argsSpace):
                        if argData[DT_ARG] is None:
                            print('No fitting dataID code found')
                            continue

                        meta = {'entity': argData[DT_ARG],
                                                      'unit' : setup['unit'],
                                                      'category':'',
                                                      'scenario' : setup['scenario'],
                                                      'source' : setup['source']}
#                        print(meta)
                        ID = core._createDatabaseID(meta)
                        
                        if ID not in tablesToReturn.keys():
                            table = Datatable()
                            table.meta = {'entity': argData[DT_ARG],
                                                      'unit' : setup['unit'],
                                                      'category':'',
                                                      'scenario' : setup['scenario'],
                                                      'source' : setup['source']}

                            tablesToReturn.add(table)
#                        print(argData)
                        value = self._readValue(wksSheet, xlsRow=argData[0], xlsCol=argData[1])

                        if isinstance(value, pd.DataFrame):
                            value = pandasStr2floatPercent(value)
                        else:
                            if value != '#VALUE!':
                                value = _str2float(value)
#                        print(value)
                        
#                        if  setup['unit'] == '%':
#                            value = value*100
#                        sdf
                        tablesToReturn[ID].loc[argData[SP_ARG],argData[TI_ARG]] = value
#                            print('success')
#                        iCount +=1
                            
#                        except Exception as e:
##                            print(e)
#                            pass
#                    self.wb.save(self.setup['fileName'])
#                            import pdb
#                            pdb.set_trace()
                        
        for table in tablesToReturn:
            table.columns = table.columns.astype(int)
            if '%' in table.meta['unit']:
                 table.loc[:,:] = table.loc[:,:]*100
                 tablesToReturn[table.ID] = table
                        
        return tablesToReturn
#            wb.save(self.setup['fileName'])
#            print('{} items inserted'.format(iCount))
#            
#            wb.close()

    def _readValue(self, wrkSheet, xlsRow, xlsCol):
#        if self.overwrite or pd.isna():
#            wrkSheet.cell(row=xlsRow, column=xlsCol, value = value)
        return wrkSheet.cell(row=xlsRow, column=xlsCol).value
#    def _writeMultipleIndicators(self, setup):
#        ws_readValues = load_workbook(self.setup['fileName'], data_only=True)[setup['sheetName']]
#        try:
#            tableIDs = [cell.value for cell in ws_readValues[setup['dataID']]]
#        except:
#            tableIDs = [cell[0].value for cell in ws_readValues[setup['dataID']]]
#            
#            tables = dt.getTables(tableIDs)
#            
#            REG_FIND_ROWS.search(setup['timeIdxList'])

class HTML_BLOCK_FILE():
    MAIN_STYLE_HEADER  = """
                <style type="text/css">
                .main-container {
                  max-width: 940px;
                  margin-left: auto;
                  margin-right: auto;
                  text-align: justify;
                }
                h1 {
                    margin-top: 40px;
                }
                code {
                  color: inherit;
                  background-color: rgba(0, 0, 0, 0.04);
                }
                img {
                  max-width:100%;
                  height: auto;
                }
                .tabbed-pane {
                  padding-top: 12px;
                }
                .html-widget {
                  margin-bottom: 20px;
                }
                button.code-folding-btn:focus {
                  outline: none;
                }
                </style>
                <div class="container-fluid main-container">
                """

    MAIN_STYLE_FOODER  = """ </div>"""

    def __init__(self, fileName, title):
        
        self.fileName = fileName
        self.headerBlocks = list()
        self.bodyBlocks   = list()
        self.fooderBlocks  = list()
        self.title = title
        self.replaceDict = dict()
        
        self.headerBlocks.append(self.MAIN_STYLE_HEADER)
        self.fooderBlocks.append(self.MAIN_STYLE_FOODER)
        
    def _writeFile(self):
        buffer = self._createBuffer()
        buffer = self._doReplacement(buffer)
        with open(self.fileName, 'w') as f:
            f.write(buffer)
            
    def write(self):
        self._writeFile()
        
    def addToHeader(self, htmlCode):
        self.headerBlocks.append(htmlCode)
    
    def addToBody(self, htmlCode):
        self.bodyBlocks.append(htmlCode)
    
    def addToFooder(self, htmlCode):
        self.fooderBlocks.append(htmlCode)

    def appendHeading(self, heading, tierStr='h1'):
        self.bodyBlocks.append("""<{tier}>{heading}</{tier}>""".format(heading=heading, tier=tierStr))
        

    def addRepacement(self, old, new):
        self.replaceDict[old] = new

    def _doReplacement(self, buffer):
        for key, value in self.replaceDict.items():
            buffer = buffer.replace(key,value)
        return buffer
    
    def _createBuffer(self):
        #%% header 
        buffer = """<html>
                <head><title>{title}</title></head>
                <link rel="stylesheet" type="text/css" href="df_style.css"/>
                """.format(title=self.title)

        for block in self.headerBlocks:
            buffer += block
            
        #body
        buffer += """<body>"""

        for block in self.bodyBlocks:
            buffer += block
            
        buffer += """</body>"""
        
        #fooder 
        for block in self.fooderBlocks:
            buffer += block        
        buffer +="""</html>"""
        return buffer
        
class HTML_File():
    
    def __init__(self, fileName):
        
        self.fileName = fileName
        self.buffer = """"""
        

        self.mainStyle_header  = """
                <style type="text/css">
                .main-container {
                  max-width: 940px;
                  margin-left: auto;
                  margin-right: auto;
                  text-align: justify;
                }
                
                code {
                  color: inherit;
                  background-color: rgba(0, 0, 0, 0.04);
                }
                img {
                  max-width:100%;
                  height: auto;
                }
                .tabbed-pane {
                  padding-top: 12px;
                }
                .html-widget {
                  margin-bottom: 20px;
                }
                button.code-folding-btn:focus {
                  outline: none;
                }
                </style>
                <div class="container-fluid main-container">
                """


        self.mainStyle_fooder  = """ </div>"""

        self._appendHeader()
        
        
    def _appendHeader(self):
        self.buffer += """
        <html>
        """ + self.mainStyle_header + """
        <head><title>HTML Pandas Dataframe with CSS</title></head>
          <link rel="stylesheet" type="text/css" href="df_style.css"/>
          <body>
        """
    
    def _appendFooder(self):
        self.buffer += self.mainStyle_fooder + """
        </body>
        </html>.
        """
        
    def _writeFile(self):
        with open(self.fileName, 'w') as f:
            f.write(self.buffer)
            

    def appendHeading(self, heading, tierStr='h1'):
        self.buffer += """<{tier}>{heading}</{tier}>""".format(heading=heading, tier=tierStr)
        
    def appendTable(self, htmlTable):
        self.buffer +=htmlTable
    
    def appendText(self, htmlText):
        self.buffer += htmlText
        
        
    def close(self):
        self._appendFooder()
        self._writeFile()
        

        
if __name__ == '__main__':
    import datatools as dt
#    setup = dict()
#    setup['filePath']  = '../../projects/NDC-analysis/03_Ready/'
#    setup['fileName']  = 'CAT_CountryAssessment_Argentina_201812.xlsx'
#    setup['sheetName'] = 'OutputWebsite'
#    setup['timeIdxList']  = ('L40', 'AZ40')
#    setup['spaceIdxList'] = ('I42', 'I42')
#    
#    ex = Extractor(setup)
#    
#    print(ex.timeIdxList)
#    timeSeries = ex.gatherData()
#    
#    ex.setup(**{'spaceIdxList': ('I42', 'I43')})
#    dataDf = ex.gatherData()
#
#    ex.setup(**{'timeIdxList': ('L40', 'L40')})
#    spatSeries = ex.gatherData()    
#
#    ex.setup(**{'spaceIdxList': ('I42', 'I42')})
#    value = ex.gatherData() 
#    
#    excelIdx2PandasIdx('AB12')    
    
    #%%
    from shutil import copyfile
    copyfile('demo_empty.xlsx', "demo.xlsx")
    ins = ExcelWriter(fileName='demo.xlsx')
    tables = ins.insert_data()
    ins.close()
    #%%    
    reader = ExcelReader_New(fileName='demo_filled.xlsx')
    out = reader.read_data()
#    for setup in ins.getSetups():
#        dataTable = dt.getTable(setup['dataID'])
#        ins._writeData(setup, dataTable)